import re
from openpyxl.utils import get_column_letter

class DataValidator:
    def __init__(self, source_loader, color_map, bz_standards):
        self.source = source_loader
        self.color_map = color_map
        self.bz_standards = bz_standards

    def check_composition(self, value, row, col, col_name):
        val = value
        if '绒子含量' in val:
            idx = val.find('绒子含量')
            percent_idx = val.find('%', idx)
            if percent_idx != -1:
                val = val[:percent_idx] + val[percent_idx + 1:]

        numbers = re.findall(r'(\d+(?:\.\d+)?)%', val)
        if not numbers:
            return None
        total = sum(float(n) for n in numbers)
        for n in numbers:
            if float(n) > 100:
                return f"第{row}行{get_column_letter(col)}列（{col_name}）含量错误【{n}%】"
        if abs(total - round(total / 100) * 100) > 0.001:
            return f"第{row}行{get_column_letter(col)}列（{col_name}）含量错误【{total:.2f}%】"
        if total > 500:
            return f"第{row}行{get_column_letter(col)}列（{col_name}）含量合计【{total:.2f}%】超过预警值500"
        return None

    def check_sensitive_words(self, value, row, col, col_name):
        base_words = ["含绒量", "羽绒棉", "棉羊毛", "山羊毛", "绵羊绒", "绵100%", "竹纤维", "牛奶丝", "大豆纤维",
                      "聚脂纤维", "聚醋纤维", "沾纤", "粘钎", "涤纶", "晴纶", "睛纶", "精纶", "尼龙", "锦伦", "绵纶", "伦",
                      "棉纶", "安纶", "氨伦", "莱卡", "莱塞尔", "来赛尔", "天丝", "莫代而", "莫代儿", "莫带尔", "木代尔", "真丝",
                      "棉麻", "人造棉", "人棉", "丝棉", "青根貂", "醋脂纤维", "PU皮", "懒兔毛", "铜氨丝", "弹力丝", "弹力纤维",
                      "冰丝", "人造丝", "丝光棉", "彩棉", "再生素纤维", "雪纺", "拉架", "玉米纤维", "仿丝棉", "金属丝",
                      "金银丝", "金丝", "银丝", "浣熊绒", "杜邦", "珊瑚绒"]
        found = []
        for w in base_words:
            if w == "绵100%":
                if re.search(r'绵\d+%', value):
                    found.append("绵X%")
            else:
                if w in value:
                    found.append(w)
        if found:
            return f"第{row}行{get_column_letter(col)}列（{col_name}）含违规词【{', '.join(found)}】"
        return None

    def check_standard_match(self, std_value):
        return std_value in self.bz_standards

    def check_price_comparison(self, row):
        std = self.source.get_numeric_cell(row, '标准价')
        pur = self.source.get_numeric_cell(row, '预估采购价')
        if std is None or pur is None:
            return None
        if std == 0 or pur == 0:
            return None
        if std < pur:
            return f"第{row}行 标准价【{std:.2f}】低于采购价【{pur:.2f}】"
        ratio = std / pur
        if ratio < 2 or ratio > 20.0:
            return f"第{row}行 标准价【{std:.2f}】与采购价【{pur:.2f}】的倍率【{ratio:.2f}】超过预警值(2-20)"
        return None

    def check_standard_price_range(self, row):
        std = self.source.get_numeric_cell(row, '标准价')
        if std is None:
            return None
        if not std.is_integer():
            return f"第{row}行 标准价【{std:.2f}】非整数"
        if std < 0 or std > 80000:
            return f"第{row}行 标准价【{std:.2f}】超过预警值8w"
        return None

    def check_purchase_price_range(self, row):
        pur = self.source.get_numeric_cell(row, '预估采购价')
        if pur is None:
            return None
        if round(pur, 2) != pur:
            return f"第{row}行 采购价【{pur:.4f}】小数位数超过2位"
        if pur < 0 or pur > 10000:
            return f"第{row}行 采购价【{pur:.2f}】超过预警值1w"
        return None

    def check_all(self, rows, global_product_count, global_style_count):
        issues = {
            'composition': [],
            'sensitive': [],
            'standard': [],
            'leather_level': [],
            'product_dup': set(),
            'style_dup': set(),
            'price_consistency': [],
            'price_compare': [],
            'standard_range': [],
            'purchase_range': [],
            'space': [],
            'colon': [],
            'origin_dup': [],
            'std_consistency': [],
            'style_rule': [],
        }

        product_origin = {}
        product_std = {}
        product_prices_std = {}
        product_prices_pur = {}

        for row in rows:
            product_key = self.source.get_cell_value(row, '品名')
            color = self.source.get_cell_value(row, '颜色')
            if color:
                product_key = f"{product_key}（{color}）"
            else:
                product_key = f"{product_key}（{' '*13}）"
            if global_product_count.get(product_key, 0) > 1:
                issues['product_dup'].add(f"品名【{product_key}】重复 {global_product_count[product_key]} 次")

            style = self.source.get_cell_value(row, '款号')
            if style and global_style_count.get(style, 0) > 1:
                issues['style_dup'].add(f"款号【{style}】重复 {global_style_count[style]} 次")

            for col_key in ['款号', '品名', '颜色', '原厂编号', '面料成分', '里料成分', '其他成分', '执行标准',
                            '安全类别', '主供应商']:
                col = self.source.col_mapper.get_col(col_key)
                if col:
                    val = self.source.cached_data.get((row, col))
                    if val is not None:
                        str_val = str(val)
                        if str_val.strip() != str_val:
                            issues['space'].append(f"第{row}行{get_column_letter(col)}列 包含前后空白字符")
                        inner_cleaned = str_val.strip()
                        if re.search(r'\s', inner_cleaned):
                            issues['space'].append(f"第{row}行{get_column_letter(col)}列 包含内部空白字符")

            for comp_key in ['面料成分', '里料成分', '其他成分']:
                col = self.source.col_mapper.get_col(comp_key)
                if col:
                    val = self.source.cached_data.get((row, col))
                    if val and ':' in str(val):
                        issues['colon'].append(f"第{row}行{get_column_letter(col)}列 含英文冒号")

            for comp_key in ['面料成分', '里料成分', '其他成分']:
                col = self.source.col_mapper.get_col(comp_key)
                if col:
                    val = self.source.cached_data.get((row, col))
                    if val:
                        res = self.check_composition(str(val), row, col, comp_key)
                        if res:
                            issues['composition'].append(res)

            for comp_key in ['面料成分', '里料成分', '其他成分']:
                col = self.source.col_mapper.get_col(comp_key)
                if col:
                    val = self.source.cached_data.get((row, col))
                    if val:
                        res = self.check_sensitive_words(str(val), row, col, comp_key)
                        if res:
                            issues['sensitive'].append(res)

            std_col = self.source.col_mapper.get_col('执行标准')
            if std_col:
                std_val = self.source.cached_data.get((row, std_col))
                if std_val:
                    if not self.check_standard_match(str(std_val).strip()):
                        issues['standard'].append(f"第{row}行{get_column_letter(std_col)}列 执行标准【{std_val}】注意确认")

            cat = self.source.get_cell_value(row, '大类')
            level = self.source.get_cell_value(row, '商品级别')
            if cat == '皮装' and level != '皮装':
                issues['leather_level'].append(f"第{row}行 大类【皮装】与商品级别【{level}】不匹配")

            res = self.check_price_comparison(row)
            if res:
                issues['price_compare'].append(res)
            res = self.check_standard_price_range(row)
            if res:
                issues['standard_range'].append(res)
            res = self.check_purchase_price_range(row)
            if res:
                issues['purchase_range'].append(res)

            prod = self.source.get_cell_value(row, '品名')
            if prod:
                origin = self.source.get_cell_value(row, '原厂编号')
                if origin:
                    product_origin.setdefault(prod, []).append(origin)
                std = self.source.get_cell_value(row, '执行标准')
                if std:
                    product_std.setdefault(prod, []).append(std)

            std_price = self.source.get_numeric_cell(row, '标准价')
            pur_price = self.source.get_numeric_cell(row, '预估采购价')
            if std_price is not None:
                product_prices_std.setdefault(prod, []).append(std_price)
            if pur_price is not None:
                product_prices_pur.setdefault(prod, []).append(pur_price)

        for prod, prices in product_prices_std.items():
            if len(prices) > 1 and len(set(prices)) > 1:
                issues['price_consistency'].append(f"品名【{prod}】存在多个标准价：{sorted(set(prices))}")
        for prod, prices in product_prices_pur.items():
            if len(prices) > 1 and len(set(prices)) > 1:
                issues['price_consistency'].append(f"品名【{prod}】存在多个采购价：{sorted(set(prices))}")

        for prod, origins in product_origin.items():
            if len(origins) > 1 and len(set(origins)) > 1:
                issues['origin_dup'].append(f"品名【{prod}】存在多个原厂编号：{sorted(set(origins))}")

        for prod, stds in product_std.items():
            if len(stds) > 1 and len(set(stds)) > 1:
                issues['std_consistency'].append(f"品名【{prod}】执行标准不一致：{sorted(set(stds))}")

        return issues