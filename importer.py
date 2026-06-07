import os
import json
import re
from openpyxl import load_workbook

class DataImporter:
    def __init__(self, source_loader, color_map, target_dir, db_manager, log_callback):
        self.source = source_loader
        self.color_map = color_map
        self.target_dir = target_dir
        self.db = db_manager
        self.log = log_callback
        self.generated_styles = {}

    def generate_style_for_row(self, row):
        product_name = self.source.get_cell_value(row, '品名')
        if not product_name:
            return None
        size_group = self.source.get_cell_value(row, '尺寸组')
        if size_group and 'B' in size_group:
            suffix = 'B'
        elif size_group and 'C' in size_group:
            suffix = 'C'
        else:
            suffix = 'A'
        color_name = self.source.get_cell_value(row, '颜色')
        if not color_name:
            return None
        norm_color = color_name.replace(' ', '').lower()
        color_code = None
        for key, code in self.color_map.items():
            if key.replace(' ', '').lower() == norm_color:
                color_code = code
                break
        if color_code is None:
            return None
        return f"{product_name}{suffix}{color_code}"

    def clean_cell_value(self, value):
        if value is None:
            return ""
        s = str(value).strip()
        s = re.sub(r'\s+', '', s)
        s = s.replace(":", "：")
        return s

    def import_to_file(self, filename, rows, history_id):
        file_path = os.path.join(self.target_dir, filename)
        if not os.path.exists(file_path):
            self.log(f"  {filename} → 文件不存在：{file_path}，已跳过")
            return False

        try:
            target_wb = load_workbook(file_path)
            target_ws = target_wb.active

            target_headers = {}
            for col in range(1, target_ws.max_column + 1):
                header = target_ws.cell(1, col).value
                if header:
                    target_headers[col] = str(header).strip()

            source_headers = self.source.col_mapper.headers

            col_map = {}
            for t_col, t_header in target_headers.items():
                matched = None
                for s_col, s_header in source_headers.items():
                    if s_header == t_header:
                        matched = s_col
                        break
                col_map[t_col] = matched

            color_target_col = None
            source_color_col = self.source.col_mapper.get_col('颜色')
            if filename == "条形生成器_批量导入.xlsx":
                for t_col, t_header in target_headers.items():
                    if t_header == "颜色":
                        color_target_col = t_col
                        break
                if color_target_col is None:
                    for t_col, t_header in target_headers.items():
                        if '颜色' in t_header and '名称' not in t_header:
                            color_target_col = t_col
                            break

            style_target_col = None
            for t_col, t_header in target_headers.items():
                if t_header == "款号":
                    style_target_col = t_col
                    break

            data_rows = []
            row_data_json_list = []
            for row in rows:
                row_dict = self.source.get_row_data_dict(row)
                row_data_json_list.append(json.dumps(row_dict, ensure_ascii=False))

                row_data = {}
                for t_col, s_col in col_map.items():
                    if t_col == color_target_col:
                        continue
                    if s_col is not None:
                        source_val = self.source.cached_data.get((row, s_col))
                    else:
                        source_val = None
                    cleaned = self.clean_cell_value(source_val) if source_val is not None else ""
                    row_data[t_col] = cleaned

                if style_target_col is not None and style_target_col not in row_data:
                    row_data[style_target_col] = ""

                if color_target_col is not None and source_color_col is not None:
                    color_name = self.source.cached_data.get((row, source_color_col))
                    if color_name:
                        color_name = str(color_name).strip()
                        norm_name = color_name.replace(' ', '').lower()
                        mapped_code = None
                        for key, code in self.color_map.items():
                            if key.replace(' ', '').lower() == norm_name:
                                mapped_code = code
                                break
                        final_val = mapped_code if mapped_code else color_name
                    else:
                        final_val = ""
                    row_data[color_target_col] = final_val
                elif color_target_col is not None:
                    row_data[color_target_col] = ""

                data_rows.append(row_data)

            if style_target_col is not None:
                for idx, row in enumerate(rows):
                    generated = self.generated_styles.get(row, "")
                    if generated:
                        data_rows[idx][style_target_col] = generated

            if target_ws.max_row >= 2:
                target_ws.delete_rows(2, target_ws.max_row - 1)

            for r_idx, row_data in enumerate(data_rows, start=2):
                for col, val in row_data.items():
                    cell = target_ws.cell(r_idx, col, val)
                    cell.number_format = '@'

            target_wb.save(file_path)
            target_wb.close()

            if history_id is not None:
                for idx, row in enumerate(rows):
                    product_name = self.source.get_cell_value(row, '品名')
                    style = self.generated_styles.get(row, "")
                    color = self.source.get_cell_value(row, '颜色')
                    status = "成功"
                    message = ""
                    if not style:
                        status = "警告"
                        message = "款号为空"
                    elif style != self.source.get_cell_value(row, '款号') and self.source.get_cell_value(row, '款号'):
                        message = "原款号被覆盖"
                    self.db.record_detail(
                        history_id, filename, row, product_name, style, color, status, message,
                        row_data_json_list[idx], import_type='data'
                    )

            return True

        except PermissionError as e:
            self.log(f"  {filename} → 保存失败：文件可能被其他程序占用 ({e})")
            return False
        except Exception as e:
            self.log(f"  {filename} → 保存失败：{e}")
            return False