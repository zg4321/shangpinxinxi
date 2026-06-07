import os
import sys
import json
import tkinter as tk
import time
from tkinter import ttk, filedialog, messagebox, simpledialog
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime, date
import re
import sqlite3
import threading


# ==================== 数据库管理 ====================
class DatabaseManager:
    def __init__(self, db_path):
        self.db_path = db_path
        self._init_db()

    def _init_db(self):
        with sqlite3.connect(self.db_path) as conn:
            cursor = conn.cursor()
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS import_history (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    import_time TIMESTAMP,
                    source_file TEXT,
                    source_sheet TEXT,
                    total_rows INTEGER,
                    target_files TEXT,
                    import_type TEXT DEFAULT 'data'
                )
            ''')
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS import_details (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    history_id INTEGER,
                    import_type TEXT DEFAULT 'data',
                    target_file TEXT,
                    source_row INTEGER,
                    product_name TEXT,
                    style TEXT,
                    color TEXT,
                    status TEXT,
                    message TEXT,
                    source_row_data TEXT,
                    FOREIGN KEY(history_id) REFERENCES import_history(id)
                )
            ''')
            # 兼容旧数据库：添加缺失字段
            cursor.execute("PRAGMA table_info(import_history)")
            columns = [col[1] for col in cursor.fetchall()]
            if "import_type" not in columns:
                cursor.execute("ALTER TABLE import_history ADD COLUMN import_type TEXT DEFAULT 'data'")
            cursor.execute("PRAGMA table_info(import_details)")
            columns = [col[1] for col in cursor.fetchall()]
            if "import_type" not in columns:
                cursor.execute("ALTER TABLE import_details ADD COLUMN import_type TEXT DEFAULT 'data'")
            if "source_row_data" not in columns:
                cursor.execute("ALTER TABLE import_details ADD COLUMN source_row_data TEXT")

    def record_history(self, source_file, source_sheet, total_rows, target_files_list, import_type='data'):
        local_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        with sqlite3.connect(self.db_path) as conn:
            cursor = conn.cursor()
            cursor.execute('''
                INSERT INTO import_history (import_time, source_file, source_sheet, total_rows, target_files, import_type)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (local_time,
                  os.path.basename(source_file),
                  source_sheet,
                  total_rows,
                  ','.join(target_files_list),
                  import_type))
            return cursor.lastrowid

    def record_detail(self, history_id, target_file, source_row, product_name, style, color, status, message, row_data_json, import_type='data'):
        with sqlite3.connect(self.db_path) as conn:
            cursor = conn.cursor()
            cursor.execute('''
                INSERT INTO import_details (history_id, import_type, target_file, source_row, product_name, style, color, status, message, source_row_data)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (history_id, import_type, target_file, source_row, product_name, style, color, status, message, row_data_json))

    def get_history(self, filters=None):
        conditions = []
        params = []
        if filters:
            if filters.get('start_date'):
                conditions.append("import_time >= ?")
                params.append(filters['start_date'])
            if filters.get('end_date'):
                conditions.append("import_time <= ?")
                params.append(filters['end_date'] + " 23:59:59")
            if filters.get('target_file') and filters['target_file'] != '全部':
                conditions.append("id IN (SELECT DISTINCT history_id FROM import_details WHERE target_file = ?)")
                params.append(filters['target_file'])
            if filters.get('import_type'):
                conditions.append("import_type = ?")
                params.append(filters['import_type'])
        where_clause = " AND ".join(conditions) if conditions else "1"
        query = f"SELECT id, import_time, source_file, source_sheet, total_rows, target_files, import_type FROM import_history WHERE {where_clause} ORDER BY import_time DESC"
        with sqlite3.connect(self.db_path) as conn:
            cursor = conn.cursor()
            cursor.execute(query, params)
            return cursor.fetchall()

    def get_details(self, history_id):
        with sqlite3.connect(self.db_path) as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT source_row, target_file, product_name, style, color, status, message, source_row_data FROM import_details WHERE history_id=?", (history_id,))
            return cursor.fetchall()

    def get_distinct_target_files(self):
        with sqlite3.connect(self.db_path) as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT DISTINCT target_file FROM import_details")
            return [row[0] for row in cursor.fetchall()]

    def delete_history(self, history_id):
        """删除指定历史记录及其关联明细（不执行 VACUUM）"""
        with sqlite3.connect(self.db_path) as conn:
            cursor = conn.cursor()
            cursor.execute("DELETE FROM import_details WHERE history_id = ?", (history_id,))
            cursor.execute("DELETE FROM import_history WHERE id = ?", (history_id,))

    def vacuum(self):
        """回收数据库空间，减小文件大小"""
        with sqlite3.connect(self.db_path) as conn:
            conn.execute("VACUUM")

# ==================== 列映射 ====================
class ColumnMapper:
    def __init__(self, ws):
        self.ws = ws
        self.headers = {}          # col -> header_str
        self.col_index = {}        # standard_key -> col
        self._build_headers()

    def _build_headers(self):
        for col in range(1, self.ws.max_column + 1):
            val = self.ws.cell(1, col).value
            self.headers[col] = str(val).strip() if val else f"列{col}"

    def map_columns(self, keyword_dict):
        for std_key, keywords in keyword_dict.items():
            for col, header in self.headers.items():
                for kw in keywords:
                    if kw in header:
                        self.col_index[std_key] = col
                        break
                if std_key in self.col_index:
                    break

    def get_col(self, std_key):
        return self.col_index.get(std_key)


# ==================== Excel 源加载器（带缓存）====================
class ExcelSourceLoader:
    def __init__(self, file_path):
        self.file_path = file_path
        self.wb = None
        self.ws = None
        self.col_mapper = None
        self.cached_data = {}          # {(row, col): value}
        self.allowed_sheets = ["帛点", "庄子", "样衣"]
        self.keyword_map = {
            '款号': ['款号'],
            '品名': ['品名'],
            '颜色': ['颜色', '(颜色)'],
            '原厂编号': ['原厂编号'],
            '主供应商': ['主供应商'],
            '商品级别': ['商品级别'],
            '大类': ['大类'],
            '中类': ['中类'],
            '子类': ['子类'],
            '内胆': ['内胆'],
            '年份': ['年份'],
            '季节': ['季节'],
            '尺寸组': ['尺寸组'],
            '品牌': ['品牌'],
            '渠道': ['渠道'],
            '退货仓': ['退货仓'],
            '预估采购价': ['预估采购价'],
            '标准价': ['标准价'],
            '面料成分': ['面料成分'],
            '里料成分': ['里料成分'],
            '其他成分': ['其他成分'],
            '执行标准': ['执行标准'],
            '安全类别': ['安全类别'],
            '备注': ['备注'],
        }
        self.load_workbook()

    def load_workbook(self):
        """加载工作簿，若已存在则先关闭旧的（释放文件句柄）"""
        if self.wb:
            try:
                self.wb.close()
            except AttributeError:
                pass  # 兼容旧版 openpyxl
        self.wb = load_workbook(self.file_path, data_only=True)

    def reload(self):
        """刷新数据：重新加载工作簿并恢复当前工作表状态"""
        self.load_workbook()
        if self.ws:
            self.set_sheet(self.ws.title)

    def close(self):
        """显式关闭工作簿，释放文件句柄"""
        if self.wb:
            try:
                self.wb.close()
            except AttributeError:
                pass
            finally:
                self.wb = None
        self.ws = None
        self.col_mapper = None
        self.cached_data.clear()

    def __del__(self):
        """析构时尝试关闭工作簿（防御性措施）"""
        self.close()

    def get_allowed_sheets(self):
        return [s for s in self.wb.sheetnames if s in self.allowed_sheets]

    def set_sheet(self, sheet_name):
        if sheet_name not in self.wb.sheetnames:
            raise ValueError(f"工作表 {sheet_name} 不存在")
        self.ws = self.wb[sheet_name]
        self.col_mapper = ColumnMapper(self.ws)
        self.col_mapper.map_columns(self.keyword_map)
        self._build_cache()

    def _build_cache(self):
        """一次性将当前工作表所有数据读入内存"""
        self.cached_data.clear()
        max_row = self.ws.max_row
        max_col = self.ws.max_column
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                val = self.ws.cell(row, col).value
                self.cached_data[(row, col)] = val

    def get_cell_value(self, row, std_key):
        col = self.col_mapper.get_col(std_key)
        if col:
            val = self.cached_data.get((row, col))
            return str(val).strip() if val is not None else ""
        return ""

    def get_numeric_cell(self, row, std_key):
        col = self.col_mapper.get_col(std_key)
        if col:
            val = self.cached_data.get((row, col))
            if val is None:
                return None
            try:
                return float(val)
            except:
                return None
        return None

    def get_row_data_dict(self, row):
        """从缓存中构建行数据字典，用于历史快照"""
        d = {}
        for col, header in self.col_mapper.headers.items():
            val = self.cached_data.get((row, col))
            if isinstance(val, datetime):
                val = val.strftime("%Y-%m-%d %H:%M:%S")
            elif val is None:
                val = ""
            else:
                val = str(val).strip()
            d[header] = val
        return d

    def load_color_map(self):
        """从CO表加载颜色名称->代码映射（直接读取工作表）"""
        color_map = {}
        if 'CO' not in self.wb.sheetnames:
            return color_map
        co_ws = self.wb['CO']
        name_col = None
        code_col = None
        for col in range(1, co_ws.max_column + 1):
            header = co_ws.cell(1, col).value
            if not header:
                continue
            h = str(header).strip()
            if '颜色名称' in h or '颜色名' in h:
                name_col = col
            elif ('颜色' in h and '名称' not in h) or '颜色代码' in h:
                code_col = col
        if name_col and code_col:
            for row in range(2, co_ws.max_row + 1):
                name = co_ws.cell(row, name_col).value
                code = co_ws.cell(row, code_col).value
                if name and code:
                    color_map[str(name).strip()] = str(code).strip()
        return color_map

    def load_bz_standards(self):
        """从BZ表加载执行标准集合"""
        standards = set()
        if 'BZ' in self.wb.sheetnames:
            bz_ws = self.wb['BZ']
            for row in range(2, bz_ws.max_row + 1):
                val = bz_ws.cell(row, 2).value
                if val:
                    standards.add(str(val).strip())
        return standards


# ==================== 数据检查器 ====================
class DataValidator:
    def __init__(self, source_loader, color_map, bz_standards):
        self.source = source_loader
        self.color_map = color_map
        self.bz_standards = bz_standards

    def check_composition(self, value, row, col, col_name):
        """成分检查 - 保留原绒子含量特殊处理"""
        val = value
        # 原逻辑：若包含'绒子含量'，删除其后的第一个百分号
        if '绒子含量' in val:
            idx = val.find('绒子含量')
            percent_idx = val.find('%', idx)
            if percent_idx != -1:
                val = val[:percent_idx] + val[percent_idx + 1:]

        numbers = re.findall(r'(\d+(?:\.\d+)?)%', val)
        if not numbers:
            return None
        total = sum(float(n) for n in numbers)
        for n in numbers:
            if float(n) > 100:
                return f"第{row}行{get_column_letter(col)}列（{col_name}）含量错误【{n}%】"
        if abs(total - round(total / 100) * 100) > 0.001:
            return f"第{row}行{get_column_letter(col)}列（{col_name}）含量错误【{total:.2f}%】"
        if total > 500:
            return f"第{row}行{get_column_letter(col)}列（{col_name}）含量合计【{total:.2f}%】超过预警值500"
        return None

    def check_sensitive_words(self, value, row, col, col_name):
        base_words = ["含绒量", "羽绒棉", "棉羊毛", "山羊毛", "绵羊绒", "绵100%", "竹纤维", "牛奶丝", "大豆纤维",
                      "聚脂纤维", "聚醋纤维", "沾纤", "粘钎", "涤纶", "晴纶", "睛纶", "精纶", "尼龙", "锦伦", "绵纶", "伦",
                      "棉纶", "安纶", "氨伦", "莱卡", "莱塞尔", "来赛尔", "天丝", "莫代而", "莫代儿", "莫带尔", "木代尔", "真丝",
                      "棉麻", "人造棉", "人棉", "丝棉", "青根貂", "醋脂纤维", "PU皮", "懒兔毛", "铜氨丝", "弹力丝", "弹力纤维",
                      "冰丝", "人造丝", "丝光棉", "彩棉", "再生素纤维", "雪纺", "拉架", "玉米纤维", "仿丝棉", "金属丝",
                      "金银丝", "金丝", "银丝", "浣熊绒", "杜邦", "珊瑚绒"]
        found = []
        for w in base_words:
            if w == "绵100%":
                if re.search(r'绵\d+%', value):
                    found.append("绵X%")
            else:
                if w in value:
                    found.append(w)
        if found:
            return f"第{row}行{get_column_letter(col)}列（{col_name}）含违规词【{', '.join(found)}】"
        return None

    def check_standard_match(self, std_value):
        return std_value in self.bz_standards

    def check_price_comparison(self, row):
        std = self.source.get_numeric_cell(row, '标准价')
        pur = self.source.get_numeric_cell(row, '预估采购价')
        if std is None or pur is None:
            return None
        if std == 0 or pur == 0:
            return None
        if std < pur:
            return f"第{row}行 标准价【{std:.2f}】低于采购价【{pur:.2f}】"
        ratio = std / pur
        if ratio < 2 or ratio > 20.0:
            return f"第{row}行 标准价【{std:.2f}】与采购价【{pur:.2f}】的倍率【{ratio:.2f}】超过预警值(2-20)"
        return None

    def check_standard_price_range(self, row):
        std = self.source.get_numeric_cell(row, '标准价')
        if std is None:
            return None
        if not std.is_integer():
            return f"第{row}行 标准价【{std:.2f}】非整数"
        if std < 0 or std > 80000:
            return f"第{row}行 标准价【{std:.2f}】超过预警值8w"
        return None

    def check_purchase_price_range(self, row):
        pur = self.source.get_numeric_cell(row, '预估采购价')
        if pur is None:
            return None
        if round(pur, 2) != pur:
            return f"第{row}行 采购价【{pur:.4f}】小数位数超过2位"
        if pur < 0 or pur > 10000:
            return f"第{row}行 采购价【{pur:.2f}】超过预警值1w"
        return None

    def check_all(self, rows, global_product_count, global_style_count):
        issues = {
            'composition': [],
            'sensitive': [],
            'standard': [],
            'leather_level': [],
            'product_dup': set(),
            'style_dup': set(),
            'price_consistency': [],
            'price_compare': [],
            'standard_range': [],
            'purchase_range': [],
            'space': [],
            'colon': [],
            'origin_dup': [],
            'std_consistency': [],
            'style_rule': [],
        }

        product_origin = {}
        product_std = {}
        product_prices_std = {}
        product_prices_pur = {}

        for row in rows:
            product_key = self.source.get_cell_value(row, '品名')
            color = self.source.get_cell_value(row, '颜色')
            if color:
                product_key = f"{product_key}（{color}）"
            else:
                product_key = f"{product_key}（{' '*13}）"
            if global_product_count.get(product_key, 0) > 1:
                issues['product_dup'].add(f"品名【{product_key}】重复 {global_product_count[product_key]} 次")

            style = self.source.get_cell_value(row, '款号')
            if style and global_style_count.get(style, 0) > 1:
                issues['style_dup'].add(f"款号【{style}】重复 {global_style_count[style]} 次")

            # 空白字符检查
            for col_key in ['款号', '品名', '颜色', '原厂编号', '面料成分', '里料成分', '其他成分', '执行标准',
                            '安全类别', '主供应商']:
                col = self.source.col_mapper.get_col(col_key)
                if col:
                    val = self.source.cached_data.get((row, col))
                    if val is not None:
                        str_val = str(val)
                        if str_val.strip() != str_val:
                            issues['space'].append(f"第{row}行{get_column_letter(col)}列 包含前后空白字符")
                        inner_cleaned = str_val.strip()
                        if re.search(r'\s', inner_cleaned):
                            issues['space'].append(f"第{row}行{get_column_letter(col)}列 包含内部空白字符")

            # 英文冒号
            for comp_key in ['面料成分', '里料成分', '其他成分']:
                col = self.source.col_mapper.get_col(comp_key)
                if col:
                    val = self.source.cached_data.get((row, col))
                    if val and ':' in str(val):
                        issues['colon'].append(f"第{row}行{get_column_letter(col)}列 含英文冒号")

            # 成分检查
            for comp_key in ['面料成分', '里料成分', '其他成分']:
                col = self.source.col_mapper.get_col(comp_key)
                if col:
                    val = self.source.cached_data.get((row, col))
                    if val:
                        res = self.check_composition(str(val), row, col, comp_key)
                        if res:
                            issues['composition'].append(res)

            # 敏感词
            for comp_key in ['面料成分', '里料成分', '其他成分']:
                col = self.source.col_mapper.get_col(comp_key)
                if col:
                    val = self.source.cached_data.get((row, col))
                    if val:
                        res = self.check_sensitive_words(str(val), row, col, comp_key)
                        if res:
                            issues['sensitive'].append(res)

            # 执行标准
            std_col = self.source.col_mapper.get_col('执行标准')
            if std_col:
                std_val = self.source.cached_data.get((row, std_col))
                if std_val:
                    if not self.check_standard_match(str(std_val).strip()):
                        issues['standard'].append(f"第{row}行{get_column_letter(std_col)}列 执行标准【{std_val}】注意确认")

            # 商品级别
            cat = self.source.get_cell_value(row, '大类')
            level = self.source.get_cell_value(row, '商品级别')
            if cat == '皮装' and level != '皮装':
                issues['leather_level'].append(f"第{row}行 大类【皮装】与商品级别【{level}】不匹配")

            # 价格比较
            res = self.check_price_comparison(row)
            if res:
                issues['price_compare'].append(res)
            res = self.check_standard_price_range(row)
            if res:
                issues['standard_range'].append(res)
            res = self.check_purchase_price_range(row)
            if res:
                issues['purchase_range'].append(res)

            # 原厂编号/执行标准一致性
            prod = self.source.get_cell_value(row, '品名')
            if prod:
                origin = self.source.get_cell_value(row, '原厂编号')
                if origin:
                    product_origin.setdefault(prod, []).append(origin)
                std = self.source.get_cell_value(row, '执行标准')
                if std:
                    product_std.setdefault(prod, []).append(std)

            # 价格一致性
            std_price = self.source.get_numeric_cell(row, '标准价')
            pur_price = self.source.get_numeric_cell(row, '预估采购价')
            if std_price is not None:
                product_prices_std.setdefault(prod, []).append(std_price)
            if pur_price is not None:
                product_prices_pur.setdefault(prod, []).append(pur_price)

        # 价格一致性汇总
        for prod, prices in product_prices_std.items():
            if len(prices) > 1 and len(set(prices)) > 1:
                issues['price_consistency'].append(f"品名【{prod}】存在多个标准价：{sorted(set(prices))}")
        for prod, prices in product_prices_pur.items():
            if len(prices) > 1 and len(set(prices)) > 1:
                issues['price_consistency'].append(f"品名【{prod}】存在多个采购价：{sorted(set(prices))}")

        for prod, origins in product_origin.items():
            if len(origins) > 1 and len(set(origins)) > 1:
                issues['origin_dup'].append(f"品名【{prod}】存在多个原厂编号：{sorted(set(origins))}")

        for prod, stds in product_std.items():
            if len(stds) > 1 and len(set(stds)) > 1:
                issues['std_consistency'].append(f"品名【{prod}】执行标准不一致：{sorted(set(stds))}")

        return issues


# ==================== 数据导入器 ====================
class DataImporter:
    def __init__(self, source_loader, color_map, target_dir, db_manager, log_callback):
        self.source = source_loader
        self.color_map = color_map
        self.target_dir = target_dir
        self.db = db_manager
        self.log = log_callback
        self.generated_styles = {}

    def generate_style_for_row(self, row):
        product_name = self.source.get_cell_value(row, '品名')
        if not product_name:
            return None
        size_group = self.source.get_cell_value(row, '尺寸组')
        if size_group and 'B' in size_group:
            suffix = 'B'
        elif size_group and 'C' in size_group:
            suffix = 'C'
        else:
            suffix = 'A'
        color_name = self.source.get_cell_value(row, '颜色')
        if not color_name:
            return None
        norm_color = color_name.replace(' ', '').lower()
        color_code = None
        for key, code in self.color_map.items():
            if key.replace(' ', '').lower() == norm_color:
                color_code = code
                break
        if color_code is None:
            return None
        return f"{product_name}{suffix}{color_code}"

    def clean_cell_value(self, value):
        if value is None:
            return ""
        s = str(value).strip()
        s = re.sub(r'\s+', '', s)
        s = s.replace(":", "：")
        return s

    def import_to_file(self, filename, rows, history_id):
        """
        将源数据导入指定的目标Excel文件（线程安全版本）
        参数:
            filename: 目标文件名
            rows: 要导入的源数据行号列表
            history_id: 历史记录ID
        返回:
            bool: 导入成功返回True，否则返回False
        """
        file_path = os.path.join(self.target_dir, filename)
        
        # 文件存在性检查（不弹窗，仅通过日志回调报告）
        if not os.path.exists(file_path):
            self.log(f"  {filename} → 文件不存在：{file_path}，已跳过")
            return False

        try:
            target_wb = load_workbook(file_path)
            target_ws = target_wb.active

            # 构建目标表头映射
            target_headers = {}
            for col in range(1, target_ws.max_column + 1):
                header = target_ws.cell(1, col).value
                if header:
                    target_headers[col] = str(header).strip()

            source_headers = self.source.col_mapper.headers

            # 建立列映射：目标列 -> 源列
            col_map = {}
            for t_col, t_header in target_headers.items():
                matched = None
                for s_col, s_header in source_headers.items():
                    if s_header == t_header:
                        matched = s_col
                        break
                col_map[t_col] = matched

            # 确定特殊列：颜色列（条形生成器专用）和款号列
            color_target_col = None
            source_color_col = self.source.col_mapper.get_col('颜色')
            if filename == "条形生成器_批量导入.xlsx":
                for t_col, t_header in target_headers.items():
                    if t_header == "颜色":
                        color_target_col = t_col
                        break
                if color_target_col is None:
                    for t_col, t_header in target_headers.items():
                        if '颜色' in t_header and '名称' not in t_header:
                            color_target_col = t_col
                            break

            style_target_col = None
            for t_col, t_header in target_headers.items():
                if t_header == "款号":
                    style_target_col = t_col
                    break

            # 准备数据行和JSON快照
            data_rows = []
            row_data_json_list = []
            for row in rows:
                # 源行完整数据快照
                row_dict = self.source.get_row_data_dict(row)
                row_data_json_list.append(json.dumps(row_dict, ensure_ascii=False))

                # 构建目标行数据
                row_data = {}
                for t_col, s_col in col_map.items():
                    if t_col == color_target_col:
                        continue  # 颜色列单独处理
                    if s_col is not None:
                        source_val = self.source.cached_data.get((row, s_col))
                    else:
                        source_val = None
                    cleaned = self.clean_cell_value(source_val) if source_val is not None else ""
                    row_data[t_col] = cleaned

                # 确保款号列存在
                if style_target_col is not None and style_target_col not in row_data:
                    row_data[style_target_col] = ""

                # 处理颜色列（条形生成器）
                if color_target_col is not None and source_color_col is not None:
                    color_name = self.source.cached_data.get((row, source_color_col))
                    if color_name:
                        color_name = str(color_name).strip()
                        norm_name = color_name.replace(' ', '').lower()
                        mapped_code = None
                        for key, code in self.color_map.items():
                            if key.replace(' ', '').lower() == norm_name:
                                mapped_code = code
                                break
                        final_val = mapped_code if mapped_code else color_name
                    else:
                        final_val = ""
                    row_data[color_target_col] = final_val
                elif color_target_col is not None:
                    row_data[color_target_col] = ""

                data_rows.append(row_data)

            # 如果存在款号列，用生成的款号覆盖
            if style_target_col is not None:
                for idx, row in enumerate(rows):
                    generated = self.generated_styles.get(row, "")
                    if generated:
                        data_rows[idx][style_target_col] = generated

            # 清空目标工作表原有数据（保留表头）
            if target_ws.max_row >= 2:
                target_ws.delete_rows(2, target_ws.max_row - 1)

            # 写入新数据
            for r_idx, row_data in enumerate(data_rows, start=2):
                for col, val in row_data.items():
                    cell = target_ws.cell(r_idx, col, val)
                    cell.number_format = '@'  # 文本格式，防止科学计数法

            target_wb.save(file_path)
            target_wb.close()

            # 记录导入明细到数据库（线程安全）
            if history_id is not None:
                for idx, row in enumerate(rows):
                    product_name = self.source.get_cell_value(row, '品名')
                    style = self.generated_styles.get(row, "")
                    color = self.source.get_cell_value(row, '颜色')
                    status = "成功"
                    message = ""
                    if not style:
                        status = "警告"
                        message = "款号为空"
                    elif style != self.source.get_cell_value(row, '款号') and self.source.get_cell_value(row, '款号'):
                        message = "原款号被覆盖"
                    self.db.record_detail(
                        history_id, filename, row, product_name, style, color, status, message,
                        row_data_json_list[idx], import_type='data'
                    )

            return True

        except PermissionError as e:
            self.log(f"  {filename} → 保存失败：文件可能被其他程序占用 ({e})")
            return False
        except Exception as e:
            self.log(f"  {filename} → 保存失败：{e}")
            return False


# ==================== 主界面类 ====================
class DataCheckerImporter:
    def __init__(self, root):
        self.root = root
        self.root.title("商品数据")
        self.root.geometry("1030x600")
        self.current_popup = None
        self.product_row_map = []

        self.config_file = os.path.join(os.path.dirname(sys.argv[0]), "config.json")
        self.config = self.load_config()
        self.config = {
            "source_file": self.config.get("source_file", ""),
            "source_sheet": self.config.get("source_sheet", "")
        }

        self.source_file_path = tk.StringVar(value=self.config.get("source_file", ""))
        self.source_sheet_name = tk.StringVar(value=self.config.get("source_sheet", ""))

        self.target_files = {
            "庄子款号_批量导入.xlsx": tk.BooleanVar(value=True),
            "帛点款号_批量导入.xlsx": tk.BooleanVar(value=True),
            "条形生成器_批量导入.xlsx": tk.BooleanVar(value=True)
        }

        self.db = DatabaseManager(os.path.join(os.path.dirname(sys.argv[0]), "import_log.db"))

        self.source_loader = None
        self.color_map = {}
        self.bz_standards = set()
        self.date_rows_map = {}
        self.row_product_key = {}
        self.all_dates = []
        self.global_product_count = {}
        self.global_style_count = {}

        self.create_widgets()
        self.center_window()
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
        self.ensure_config_fields()

        if self.source_file_path.get() and os.path.exists(self.source_file_path.get()):
            self.load_sheets_from_config()

    # -------------------- 界面构建 --------------------
    def create_widgets(self):
        frame_source = ttk.LabelFrame(self.root, text="源数据", padding=5)
        frame_source.pack(fill=tk.X, padx=5, pady=5)

        ttk.Label(frame_source, text="源文件:").grid(row=0, column=0, sticky=tk.W, padx=2)
        ttk.Entry(frame_source, textvariable=self.source_file_path, width=50).grid(row=0, column=1, padx=2)
        ttk.Button(frame_source, text="浏览...", command=self.browse_source_file).grid(row=0, column=2, padx=2)

        ttk.Label(frame_source, text="源工作表:").grid(row=1, column=0, sticky=tk.W, padx=2, pady=5)
        self.sheet_combo = ttk.Combobox(frame_source, textvariable=self.source_sheet_name, state="readonly", width=30)
        self.sheet_combo.grid(row=1, column=1, sticky=tk.W, padx=2, pady=5)
        self.sheet_combo.bind("<<ComboboxSelected>>", self.on_sheet_selected)
        ttk.Button(frame_source, text="刷新", command=self.refresh_source).grid(row=1, column=2, padx=2)

        # 右侧放置两个按钮：导入记录、功能说明（靠右）
        frame_source.grid_columnconfigure(3, weight=1)   # 占位列，将后面的按钮推到右侧
        ttk.Button(frame_source, text="导入记录", command=self.show_import_history).grid(row=1, column=4, padx=2, sticky=tk.E)
        ttk.Button(frame_source, text="功能说明", command=self.show_help).grid(row=1, column=5, padx=2, sticky=tk.E)

        frame_middle = ttk.Frame(self.root)
        frame_middle.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        frame_date = ttk.LabelFrame(frame_middle, text="日期（可多选）", width=160)
        frame_date.pack(side=tk.LEFT, fill=tk.BOTH, expand=False)
        frame_date.pack_propagate(False)
        self.date_listbox = tk.Listbox(frame_date, selectmode=tk.EXTENDED, exportselection=False)
        self.date_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        date_scroll = ttk.Scrollbar(frame_date, orient=tk.VERTICAL, command=self.date_listbox.yview)
        date_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.date_listbox.config(yscrollcommand=date_scroll.set)
        self.date_listbox.bind("<<ListboxSelect>>", self.on_date_select)

        frame_product = ttk.LabelFrame(frame_middle, text="行号∣    品名（颜色）可多选", width=200)
        frame_product.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        frame_product.pack_propagate(False)
        self.product_listbox = tk.Listbox(frame_product, selectmode=tk.EXTENDED, exportselection=False)
        self.product_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        product_scroll = ttk.Scrollbar(frame_product, orient=tk.VERTICAL, command=self.product_listbox.yview)
        product_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.product_listbox.config(yscrollcommand=product_scroll.set)
        self.product_listbox.bind("<Double-Button-1>", self.on_product_double_click)
        self.product_listbox.bind('<KeyPress-Up>', self.on_product_key_up_down)
        self.product_listbox.bind('<KeyPress-Down>', self.on_product_key_up_down)
        self.product_listbox.bind('<ButtonRelease-1>', self.on_product_click)

        frame_log = ttk.LabelFrame(frame_middle, text="日志", width=500)
        frame_log.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        frame_log.pack_propagate(False)
        self.log_text = tk.Text(frame_log, wrap=tk.WORD, state=tk.DISABLED)
        self.log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        log_scroll = ttk.Scrollbar(frame_log, orient=tk.VERTICAL, command=self.log_text.yview)
        log_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.log_text.config(yscrollcommand=log_scroll.set)
        try:
            self.log_text.config(font=("微软雅黑", 9))
        except:
            pass

        frame_target = ttk.LabelFrame(self.root, text="选择目标工作簿", padding=5)
        frame_target.pack(fill=tk.X, padx=5, pady=5)
        for i, (name, var) in enumerate(self.target_files.items()):
            ttk.Checkbutton(frame_target, text=name, variable=var).grid(row=0, column=i, padx=10, sticky=tk.W)

        self.progress = ttk.Progressbar(self.root, mode='determinate')
        self.progress.pack(fill=tk.X, padx=5, pady=5)
        self.progress.pack_forget()

        frame_buttons = ttk.Frame(self.root)
        frame_buttons.pack(fill=tk.X, padx=5, pady=20)

        ttk.Button(frame_buttons, text="数据检查", command=self.check_data).pack(side=tk.LEFT, padx=20)
        self.btn_import = ttk.Button(frame_buttons, text="数据导入", command=self.import_data)
        self.btn_import.pack(side=tk.LEFT, padx=0)
        ttk.Button(frame_buttons, text="颜色导入", command=self.show_color_import_dialog).pack(side=tk.LEFT, padx=20)
        ttk.Button(frame_buttons, text="供应商导入", command=self.show_supplier_import_dialog).pack(side=tk.LEFT, padx=0)
        ttk.Button(frame_buttons, text="全选/取消", command=self.toggle_select_all).pack(side=tk.LEFT, padx=40)
        ttk.Button(frame_buttons, text="清空日志", command=self.clear_log).pack(side=tk.LEFT, padx=0)
        ttk.Button(frame_buttons, text="退出", command=self.root.quit).pack(side=tk.RIGHT, padx=20)

        self.set_widgets_state(tk.DISABLED)

    def set_widgets_state(self, state):
        self.date_listbox.config(state=state)
        self.product_listbox.config(state=state)
        for child in self.root.winfo_children():
            if isinstance(child, ttk.Frame):
                for btn in child.winfo_children():
                    if isinstance(btn, ttk.Button) and btn['text'] not in ['浏览', '功能说明', '退出', '导入记录']:
                        btn.config(state=state)

    def center_window(self):
        self.root.update_idletasks()
        x = (self.root.winfo_screenwidth() - 1030) // 4
        y = (self.root.winfo_screenheight() - 600) // 4
        self.root.geometry(f"1030x600+{x}+{y}")

    # -------------------- 配置 --------------------
    def load_config(self):
        if os.path.exists(self.config_file):
            try:
                with open(self.config_file, "r", encoding="utf-8") as f:
                    return json.load(f)
            except:
                return {}
        return {}

    def save_config(self):
        self.config = {
            "source_file": self.source_file_path.get(),
            "source_sheet": self.source_sheet_name.get()
        }
        try:
            with open(self.config_file, "w", encoding="utf-8") as f:
                json.dump(self.config, f, ensure_ascii=False, indent=2)
            self.log("配置文件已保存")
        except Exception as e:
            self.log(f"保存配置文件失败：{e}")

    def ensure_config_fields(self):
        modified = False
        if "source_file" not in self.config:
            self.config["source_file"] = ""
            modified = True
        if "source_sheet" not in self.config:
            self.config["source_sheet"] = ""
            modified = True
        if modified:
            self.save_config()

    # -------------------- 源文件操作 --------------------
    def browse_source_file(self):
        file_path = filedialog.askopenfilename(
            title="选择源Excel文件",
            filetypes=[("Excel files", "*.xlsx *.xlsm"), ("All files", "*.*")]
        )
        if file_path:
            self.source_file_path.set(file_path)
            self.load_sheets()

    def load_sheets(self):
        if self.source_loader:
            self.source_loader.close()       
        
        try:
            self.source_loader = ExcelSourceLoader(self.source_file_path.get())
            allowed = self.source_loader.get_allowed_sheets()
            self.sheet_combo['values'] = allowed
            if allowed:
                if self.source_sheet_name.get() in allowed:
                    self.sheet_combo.current(allowed.index(self.source_sheet_name.get()))
                else:
                    self.sheet_combo.current(0)
                    self.source_sheet_name.set(allowed[0])
                self.on_sheet_selected()
            else:
                self.source_sheet_name.set("")
                self.sheet_combo.set("")
                self.set_widgets_state(tk.DISABLED)
                self.log("工作簿中未找到允许的工作表（庄子、样衣、帛点）")
        except Exception as e:
            messagebox.showerror("错误", f"无法打开文件：{e}")

    def load_sheets_from_config(self):
        if self.source_loader:
            self.source_loader.close()
        try:
            self.source_loader = ExcelSourceLoader(self.source_file_path.get())
            allowed = self.source_loader.get_allowed_sheets()
            self.sheet_combo['values'] = allowed
            if allowed:
                if self.source_sheet_name.get() in allowed:
                    self.sheet_combo.current(allowed.index(self.source_sheet_name.get()))
                    self.on_sheet_selected()
                else:
                    self.sheet_combo.current(0)
                    self.source_sheet_name.set(allowed[0])
                    self.on_sheet_selected()
            else:
                self.source_sheet_name.set("")
                self.set_widgets_state(tk.DISABLED)
                self.log("工作簿中未找到允许的工作表（庄子、样衣、帛点）")
        except Exception as e:
            self.log(f"自动加载失败：{e}")
            self.set_widgets_state(tk.DISABLED)

    def refresh_source(self):
        self._clear_log_content()
        if self.source_loader:
            try:
                self.source_loader.reload()
                allowed = self.source_loader.get_allowed_sheets()
                self.sheet_combo['values'] = allowed
                if self.source_sheet_name.get() in allowed:
                    self.source_loader.set_sheet(self.source_sheet_name.get())
                    self.load_data_from_sheet()
                    self.log("源工作表数据已刷新！")
                else:
                    if allowed:
                        self.sheet_combo.current(0)
                        self.source_sheet_name.set(allowed[0])
                        self.on_sheet_selected()
                    else:
                        self.source_sheet_name.set("")
                        self.set_widgets_state(tk.DISABLED)
                        self.log("刷新后未找到允许的工作表")
            except Exception as e:
                messagebox.showerror("错误", f"刷新失败：{e}")

    def on_sheet_selected(self, event=None):
        self._clear_log_content()
        sheet = self.source_sheet_name.get()
        if not sheet:
            return
        try:
            self.source_loader.set_sheet(sheet)
            self.load_data_from_sheet()
            self.set_widgets_state(tk.NORMAL)
            self.save_config()
        except Exception as e:
            messagebox.showerror("错误", f"加载工作表失败：{e}")

    def load_data_from_sheet(self):
        ws = self.source_loader.ws
        self.date_rows_map.clear()
        self.row_product_key.clear()
        all_dates_set = set()
        for row in range(2, ws.max_row + 1):
            date_cell = self.source_loader.cached_data.get((row, 1))
            if date_cell:
                date_str = self.parse_date(date_cell)
                if not date_str:
                    continue
                product_name = self.source_loader.get_cell_value(row, '品名')
                if not product_name:
                    continue
                color_name = self.source_loader.get_cell_value(row, '颜色') or ""
                product_key = f"{product_name}（{color_name}）" if color_name else f"{product_name}（{' '*13}）"
                if date_str not in self.date_rows_map:
                    self.date_rows_map[date_str] = []
                self.date_rows_map[date_str].append(row)
                self.row_product_key[row] = product_key
                all_dates_set.add(date_str)

        self.all_dates = sorted(all_dates_set, key=lambda d: datetime.strptime(d, "%Y-%m-%d"), reverse=True)
        self.date_listbox.delete(0, tk.END)
        for d in self.all_dates:
            self.date_listbox.insert(tk.END, d)
        if self.all_dates:
            self.date_listbox.selection_set(0)
            self.update_product_list()
        self.log("数据加载完成！")

    def parse_date(self, date_val):
        if isinstance(date_val, datetime):
            return date_val.strftime("%Y-%m-%d")
        elif isinstance(date_val, str):
            for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y", "%m/%d/%Y"):
                try:
                    return datetime.strptime(date_val, fmt).strftime("%Y-%m-%d")
                except:
                    continue
        return None

    def on_date_select(self, event=None):
        self.update_product_list()

    def update_product_list(self):
        self.product_listbox.delete(0, tk.END)
        self.product_row_map = []
        selected_indices = self.date_listbox.curselection()
        if not selected_indices:
            return
        selected_dates = [self.date_listbox.get(i) for i in selected_indices]
        rows = []
        for date in selected_dates:
            rows.extend(self.date_rows_map.get(date, []))
        rows = sorted(set(rows))
        for row in rows:
            product_key = self.row_product_key.get(row, "未知")
            display_text = f"{row:>4} ∣    {product_key}"
            self.product_listbox.insert(tk.END, display_text)
            self.product_row_map.append(row)
        self.select_all_products()

    def select_all_products(self):
        self.product_listbox.selection_set(0, tk.END)

    def get_selected_rows(self):
        date_indices = self.date_listbox.curselection()
        if not date_indices:
            return []
        selected_dates = [self.date_listbox.get(i) for i in date_indices]
        all_rows = set()
        for d in selected_dates:
            all_rows.update(self.date_rows_map.get(d, []))
        prod_indices = self.product_listbox.curselection()
        if prod_indices:
            selected_rows = [self.product_row_map[i] for i in prod_indices]
            final_rows = [r for r in selected_rows if r in all_rows]
        else:
            final_rows = []
        return final_rows

    def toggle_select_all(self):
        items_count = self.product_listbox.size()
        selected = self.product_listbox.curselection()
        if len(selected) == items_count:
            self.product_listbox.selection_clear(0, tk.END)
        else:
            self.product_listbox.selection_set(0, tk.END)

    # -------------------- 双击查看数据 --------------------
    def on_product_double_click(self, event):
        selection = self.product_listbox.curselection()
        if not selection:
            return
        idx = selection[0]
        row = self.product_row_map[idx]
        if self.current_popup and self.current_popup.winfo_exists():
            self.update_popup_content(row)
        else:
            self.create_popup(row)

    def on_product_key_up_down(self, event):
        self.product_listbox.after(10, self.update_popup_for_current_selection)

    def on_product_click(self, event):
        self.update_popup_for_current_selection()

    def update_popup_for_current_selection(self):
        selection = self.product_listbox.curselection()
        if selection and self.current_popup and self.current_popup.winfo_exists():
            idx = selection[0]
            row = self.product_row_map[idx]
            self.update_popup_content(row)

    def create_popup(self, row):
        top = tk.Toplevel(self.root)
        top.title(f"查看数据 - 第 {row} 行")
        top.geometry("600x700")
        top.resizable(True, True)
        top.update_idletasks()
        x = self.root.winfo_x() + self.root.winfo_width() // 2 - 50
        y = self.root.winfo_y() + self.root.winfo_height() // 2 - 310
        top.geometry(f"+{x}+{y}")
        top.transient(self.root)

        main_frame = ttk.Frame(top)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        canvas = tk.Canvas(main_frame, highlightthickness=0)
        v_scrollbar = ttk.Scrollbar(main_frame, orient=tk.VERTICAL, command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        window_id = canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        
        # 关键修改：当canvas大小改变时，调整内部frame的宽度，使其填满canvas
        def on_canvas_configure(event):
            canvas.itemconfig(window_id, width=event.width)  # 宽度等于canvas宽度
        canvas.bind("<Configure>", on_canvas_configure)
        
        canvas.configure(yscrollcommand=v_scrollbar.set)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        v_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        entries = {}
        # 为获得更美观的布局，使用grid管理每行，让entry占据剩余宽度
        for col_key, col_idx in self.source_loader.col_mapper.col_index.items():
            cell_value = self.source_loader.cached_data.get((row, col_idx))
            display_value = str(cell_value) if cell_value is not None else ""
            
            frame_row = ttk.Frame(scrollable_frame)
            frame_row.pack(fill=tk.X, pady=2, padx=5)
            frame_row.columnconfigure(1, weight=1)  # 第1列（entry）可扩展
            
            label = ttk.Label(frame_row, text=f"{col_key}:", width=12, anchor=tk.E)
            label.grid(row=0, column=0, sticky=tk.W, padx=(0,5))
            
            entry = ttk.Entry(frame_row)
            entry.insert(0, display_value)
            entry.config(state='readonly')
            entry.grid(row=0, column=1, sticky=tk.EW)  # 横向填充
            
            entries[col_key] = (col_idx, entry)

        btn_frame = ttk.Frame(top)
        btn_frame.pack(fill=tk.X, padx=5, pady=5)

        def close_popup():
            self.current_popup = None
            top.destroy()
        ttk.Button(btn_frame, text="关闭", command=close_popup).pack(side=tk.RIGHT, padx=30)
        
        top.entries = entries
        top.row = row
        top.protocol("WM_DELETE_WINDOW", close_popup)
        self.current_popup = top
    
    def update_popup_content(self, row):
        popup = self.current_popup
        if not popup or not popup.winfo_exists():
            return
        popup.title(f"查看数据 - 第 {row} 行")
        for col_key, (col_idx, entry) in popup.entries.items():
            entry.config(state='normal')
            cell_value = self.source_loader.cached_data.get((row, col_idx))
            display_value = str(cell_value) if cell_value is not None else ""
            entry.delete(0, tk.END)
            entry.insert(0, display_value)
            entry.config(state='readonly')
        popup.row = row
        
    def on_closing(self):
        """程序退出前释放 Excel 文件句柄"""
        if self.source_loader:
            self.source_loader.close()
        self.root.destroy()

    # -------------------- 数据检查 --------------------
    def check_data(self):
        self._clear_log_content()
        rows = self.get_selected_rows()
        if not rows:
            self.log("错误：请至少选择一个日期和品名")
            return
        self.log("开始数据检查...")
        self.log("")
        self.color_map = self.source_loader.load_color_map()
        self.bz_standards = self.source_loader.load_bz_standards()
        validator = DataValidator(self.source_loader, self.color_map, self.bz_standards)

        self.scan_global_duplicates()
        issues = validator.check_all(rows, self.global_product_count, self.global_style_count)

        total_issues = 0
        category_map = {
            'composition': '成分含量',
            'sensitive': '成分违规词',
            'standard': '执行标准',
            'product_dup': '品名重复',
            'style_dup': '款号重复',
            'style_rule': '款号规则',
            'origin_dup': '原厂编号不一致',
            'std_consistency': '执行标准不一致',
            'price_consistency': '同款价格不一致',
            'price_compare': '价格比较',
            'standard_range': '标准价范围',
            'purchase_range': '采购价范围',
            'leather_level': '商品级别',
            'space': '内容含空白字符',
            'colon': '英文冒号',
        }
        categories = []
        for key, title in category_map.items():
            items = list(issues[key]) if key in ['product_dup', 'style_dup'] else issues[key]
            if items:
                categories.append(title)
                self.log(f"----【{title}】----")
                for item in items:
                    if not item.startswith("异常提示："):
                        item = "异常提示：" + item
                    self.log(item)
                    total_issues += 1
        self.log("")
        self.log("数据检查完成！")
        self.log(f"共检查行数：{len(rows)}")
        self.log(f"发现异常提示：{total_issues}")
        self.last_check_issues = total_issues
        self.last_check_categories = categories

        color_names = set()
        for row in rows:
            color_name = self.source_loader.get_cell_value(row, '颜色')
            if color_name:
                color_names.add(color_name.strip())
        missing = [c for c in color_names if c not in self.color_map]
        if missing:
            color_list = "\n".join(missing)
            if messagebox.askyesno("颜色缺失", f"以下颜色在CO表中不存在：\n{color_list}\n\n是否处理缺失颜色？"):
                new_mappings = self._prompt_missing_colors(missing)
                if new_mappings:
                    self.color_map.update(new_mappings)
                    self.log(f"已为本次会话启用 {len(new_mappings)} 个临时颜色代码（未写入CO表）。")
            else:
                self.log("用户未处理缺失颜色。")

    def scan_global_duplicates(self):
        self.global_product_count.clear()
        self.global_style_count.clear()
        for row, product_key in self.row_product_key.items():
            self.global_product_count[product_key] = self.global_product_count.get(product_key, 0) + 1
            style = self.source_loader.get_cell_value(row, '款号')
            if style:
                self.global_style_count[style] = self.global_style_count.get(style, 0) + 1

    # -------------------- 数据导入（后台线程）--------------------
    def import_data(self):
        rows = self.get_selected_rows()
        if not rows:
            self.log("错误：请至少选择一个日期和品名")
            return
        if not messagebox.askyesno("确认导入", "是否确认导入选中的数据？"):
            return

        # 检查价格
        missing_prices = []
        for row in rows:
            std = self.source_loader.get_numeric_cell(row, '标准价')
            pur = self.source_loader.get_numeric_cell(row, '预估采购价')
            if std is None:
                missing_prices.append((row, '标准价'))
            if pur is None:
                missing_prices.append((row, '预估采购价'))
        if missing_prices:
            msg_lines = ["以下行价格为空，无法导入："]
            for row, price_name in missing_prices:
                msg_lines.append(f"第{row}行 {price_name}")
            messagebox.showerror("导入错误", "\n".join(msg_lines))
            return

        only_bar = (not self.target_files['庄子款号_批量导入.xlsx'].get() and
                    not self.target_files['帛点款号_批量导入.xlsx'].get() and
                    self.target_files['条形生成器_批量导入.xlsx'].get())
        if not only_bar:
            self.check_data()
            if hasattr(self, 'last_check_issues') and self.last_check_issues > 0:
                cat_str = "\n".join(self.last_check_categories)
                if not messagebox.askyesno("数据异常", f"存在以下异常：\n{cat_str}\n\n是否忽略并继续导入？"):
                    return

        self.log("")
        self.log("开始导入数据...")

        self.color_map = self.source_loader.load_color_map()
        importer = DataImporter(self.source_loader, self.color_map,
                                os.path.dirname(self.source_file_path.get()),
                                self.db, self.log)
        for row in rows:
            actual_style = self.source_loader.get_cell_value(row, '款号')
            if actual_style and actual_style.strip():
                importer.generated_styles[row] = actual_style.strip()
            else:
                gen = importer.generate_style_for_row(row)
                if gen:
                    importer.generated_styles[row] = gen
                    self.log(f"第{row}行 款号为空，已自动生成款号：{gen}")
                else:
                    importer.generated_styles[row] = ""
        selected_files = [name for name, var in self.target_files.items() if var.get()]
        if not selected_files:
            self.log("未选择任何目标文件")
            return

        # 检查目标文件是否存在（必须在主线程弹出对话框）
        for name in selected_files:
            file_path = os.path.join(os.path.dirname(self.source_file_path.get()), name)
            if not os.path.exists(file_path):
                ans = messagebox.askyesno("文件不存在", f"在源文件目录找不到：{name}\n是否手动选择？")
                if ans:
                    new_path = filedialog.askopenfilename(
                        title=f"请选择{name}",
                        initialdir=os.path.dirname(self.source_file_path.get()),
                        filetypes=[("Excel files", "*.xlsx")]
                    )
                    if not new_path:
                        self.log(f"跳过：{name}")
                        selected_files.remove(name)
                    else:
                        # 注意：这里只是记录了路径，实际导入时仍使用原路径，需要处理。
                        # 为简化，此处先取消导入，提示用户手动放置文件。
                        messagebox.showwarning("警告", "手动选择功能暂未实现，请将目标文件放置在源文件目录下。")
                        return
                else:
                    self.log(f"跳过：{name}")
                    selected_files.remove(name)
        if not selected_files:
            return

        history_id = self.db.record_history(self.source_file_path.get(),
                                            self.source_sheet_name.get(),
                                            len(rows), selected_files, import_type='data')

        # 进度条初始化
        self.progress.config(mode='determinate', maximum=len(selected_files), value=0)
        self.progress.pack(fill=tk.X, padx=5, pady=5)
        self.btn_import.config(state=tk.DISABLED)

        thread = threading.Thread(target=self._import_thread,
                                  args=(importer, selected_files, rows, history_id),
                                  daemon=True)
        thread.start()

    def _import_thread(self, importer, selected_files, rows, history_id):
        total = len(selected_files)
        success_count = 0
        for idx, name in enumerate(selected_files):
            if importer.import_to_file(name, rows, history_id):
                success_count += 1
            self.root.after(0, self._update_import_progress, idx + 1, total)
        self.root.after(0, self._import_finished, success_count)

    def _update_import_progress(self, current, total):
        self.progress['value'] = current
        self.progress.update_idletasks()
        self.log(f"导入进度: {current}/{total} 个文件")

    def _import_finished(self, success_count):
        self.progress.stop()
        self.progress.pack_forget()
        self.btn_import.config(state=tk.NORMAL)
        self.log("")
        self.log("数据导入完成！")
        self.log(f"成功导入到 {success_count} 个模板文件！")

    # -------------------- 颜色导入 --------------------
    def show_color_import_dialog(self):
        if not self.source_loader:
            messagebox.showerror("错误", "请先加载源文件")
            return
        if 'CO' not in self.source_loader.wb.sheetnames:
            messagebox.showerror("错误", "源文件中未找到'CO'工作表")
            return

        co_ws = self.source_loader.wb['CO']
        code_col = None
        name_col = None
        for col in range(1, co_ws.max_column + 1):
            header = co_ws.cell(1, col).value
            if not header:
                continue
            h = str(header).strip()
            if '颜色名称' in h or '颜色名' in h:
                name_col = col
            elif ('颜色' in h and '名称' not in h) or '颜色代码' in h:
                code_col = col
        if not name_col or not code_col:
            messagebox.showerror("错误", "CO工作表中未找到颜色名称列或颜色代码列")
            return

        color_list = []
        for row in range(2, co_ws.max_row + 1):
            code = co_ws.cell(row, code_col).value
            name = co_ws.cell(row, name_col).value
            if code and name:
                color_list.append((str(code).strip(), str(name).strip()))
        if not color_list:
            messagebox.showinfo("提示", "CO工作表中没有有效的颜色数据")
            return

        dialog = tk.Toplevel(self.root)
        dialog.title("选择要导入的颜色")
        dialog.geometry("300x450")
        dialog.transient(self.root)
        dialog.grab_set()

        frame = ttk.Frame(dialog)
        frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        scrollbar = ttk.Scrollbar(frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        listbox = tk.Listbox(frame, selectmode=tk.EXTENDED, yscrollcommand=scrollbar.set, height=20)
        listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=listbox.yview)
        for code, name in color_list:
            listbox.insert(tk.END, f"{code}  {name}")
        if color_list:
            listbox.selection_set(len(color_list)-1)
            listbox.see(len(color_list)-1)

        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(fill=tk.X, padx=5, pady=5)

        def select_all():
            listbox.selection_set(0, tk.END)
        def clear_selection():
            listbox.selection_clear(0, tk.END)
        def import_selected():
            selected = listbox.curselection()
            if not selected:
                messagebox.showwarning("警告", "请至少选择一个颜色", parent=dialog)
                return
            items = [color_list[i] for i in selected]
            dialog.destroy()
            # 改为显示预览窗口，不再直接导入
            self.show_color_import_preview(items)

        ttk.Button(btn_frame, text="导入", command=import_selected).pack(side=tk.LEFT, padx=20)
        ttk.Button(btn_frame, text="关闭", command=dialog.destroy).pack(side=tk.LEFT, padx=5)

    def show_color_import_preview(self, color_items):
        """颜色导入预览窗口：左右两个区域（颜色定义、属性维护），每个区域内上下布局（新数据在上，旧数据在下），旧数据灰色背景"""
        source_dir = os.path.dirname(self.source_file_path.get())
        color_def_file = "颜色定义_批量导入.xlsx"
        attr_file = "属性维护_批量导入.xlsx"
        color_def_path = os.path.join(source_dir, color_def_file)
        attr_path = os.path.join(source_dir, attr_file)

        # ---------- 定义表格样式（背景色区分）----------
        style = ttk.Style()
        style.configure("ColorNew.Treeview", background="white", fieldbackground="white")
        style.configure("ColorOld.Treeview",
                        background="#f0f0f0",
                        fieldbackground="#f0f0f0",
                        foreground="black")
        style.map("ColorOld.Treeview",
                  background=[('selected', '#c0c0c0')],
                  fieldbackground=[('selected', '#c0c0c0')])
        style.configure("ColorOld.Treeview.Heading", background="#dcdcdc", foreground="black")
        style.configure("ColorOld.Vertical.TScrollbar", troughcolor="#f0f0f0", background="#d0d0d0")
        style.configure("ColorOld.Horizontal.TScrollbar", troughcolor="#f0f0f0", background="#d0d0d0")

        # ---------- 读取现有数据 ----------
        def load_existing_data(file_path, expected_headers):
            if not os.path.exists(file_path):
                return [], []
            try:
                wb = load_workbook(file_path)
                ws = wb.active
                existing_headers = []
                for col in range(1, ws.max_column + 1):
                    val = ws.cell(1, col).value
                    existing_headers.append(str(val).strip() if val else f"列{col}")
                header_to_col = {h: idx for idx, h in enumerate(existing_headers)}
                data_rows = []
                for row in range(2, ws.max_row + 1):
                    row_data = []
                    for h in expected_headers:
                        if h in header_to_col:
                            col_idx = header_to_col[h] + 1
                            val = ws.cell(row, col_idx).value
                            row_data.append(str(val).strip() if val is not None else "")
                        else:
                            row_data.append("")
                    if any(row_data):
                        data_rows.append(row_data)
                wb.close()
                return existing_headers, data_rows
            except Exception as e:
                self.log(f"读取现有数据失败 {file_path}: {e}")
                return [], []

        # 构建新数据
        color_def_new = [[code, name, ""] for code, name in color_items]      # 颜色, 颜色名称, 备注
        attr_new = [["(颜色)", f"Y{code}", name] for code, name in color_items]  # 属性定义, 属性代码, 属性描述

        # 读取现有数据
        color_def_headers, color_def_old = load_existing_data(color_def_path, ["颜色", "颜色名称", "备注"])
        attr_headers, attr_old = load_existing_data(attr_path, ["属性定义", "属性代码", "属性描述"])

        # 创建预览窗口
        dialog = tk.Toplevel(self.root)
        dialog.title("颜色导入预览 - 最终确认")
        dialog.geometry("1000x750")  # 加宽以容纳左右两个区域
        dialog.transient(self.root)
        dialog.grab_set()

        ttk.Label(dialog, text=f"即将导入 {len(color_items)} 条颜色数据，请确认：\n"
                               f"  • 颜色定义_批量导入.xlsx —— 覆盖\"颜色\"、\"颜色名称\"列，备注保留原值？\n"
                               f"  • 属性维护_批量导入.xlsx —— 覆盖\"属性定义\"、\"属性代码\"、\"属性描述\"列\n"
                               f"左右两侧分别展示两个文件将写入的数据（新旧对比）。",
                  font=("微软雅黑", 10)).pack(pady=10)

        # 左右分割主窗口
        main_paned = ttk.PanedWindow(dialog, orient=tk.HORIZONTAL)
        main_paned.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # ==================== 左侧：颜色定义 ====================
        left_frame = ttk.Frame(main_paned)
        main_paned.add(left_frame, weight=1)
        paned_left = ttk.PanedWindow(left_frame, orient=tk.VERTICAL)
        paned_left.pack(fill=tk.BOTH, expand=True)

        # 上部：新数据
        top_left = ttk.Frame(paned_left)
        paned_left.add(top_left, weight=1)
        ttk.Label(top_left, text="▼ 颜色定义 - 即将导入的新数据", font=("微软雅黑", 9, "bold")).pack(anchor=tk.W, padx=5, pady=(5,0))
        tree_new_def = ttk.Treeview(top_left, columns=("col1","col2","col3"), show="headings", height=6, style="ColorNew.Treeview")
        tree_new_def.heading("col1", text="颜色")
        tree_new_def.heading("col2", text="颜色名称")
        tree_new_def.heading("col3", text="备注")
        tree_new_def.column("col1", width=100, anchor=tk.CENTER)
        tree_new_def.column("col2", width=150, anchor=tk.CENTER)
        tree_new_def.column("col3", width=150, anchor=tk.CENTER)
        for row in color_def_new:
            tree_new_def.insert("", tk.END, values=row)
        vsb_new_def = ttk.Scrollbar(top_left, orient=tk.VERTICAL, command=tree_new_def.yview)
        hsb_new_def = ttk.Scrollbar(top_left, orient=tk.HORIZONTAL, command=tree_new_def.xview)
        tree_new_def.configure(yscrollcommand=vsb_new_def.set, xscrollcommand=hsb_new_def.set)
        tree_new_def.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
        vsb_new_def.pack(side=tk.RIGHT, fill=tk.Y)
        hsb_new_def.pack(side=tk.BOTTOM, fill=tk.X)

        # 下部：旧数据（灰色背景）
        bottom_left = tk.Frame(paned_left, bg='#f0f0f0')
        paned_left.add(bottom_left, weight=1)
        tk.Label(bottom_left, text="▼ 目标文件中已有数据（只读）", font=("微软雅黑", 9, "bold"), bg='#f0f0f0').pack(anchor=tk.W, padx=5, pady=(5,0))
        table_left = tk.Frame(bottom_left, bg='#f0f0f0')
        table_left.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        tree_old_def = ttk.Treeview(table_left, columns=("col1","col2","col3"), show="headings", height=6, style="ColorOld.Treeview")
        tree_old_def.heading("col1", text="颜色")
        tree_old_def.heading("col2", text="颜色名称")
        tree_old_def.heading("col3", text="备注")
        tree_old_def.column("col1", width=100, anchor=tk.CENTER)
        tree_old_def.column("col2", width=150, anchor=tk.CENTER)
        tree_old_def.column("col3", width=150, anchor=tk.CENTER)
        for row in color_def_old:
            tree_old_def.insert("", tk.END, values=row)
        if not color_def_old:
            tree_old_def.insert("", tk.END, values=["(无已有数据)", "", ""])
        vsb_old_def = ttk.Scrollbar(table_left, orient=tk.VERTICAL, command=tree_old_def.yview, style="ColorOld.Vertical.TScrollbar")
        hsb_old_def = ttk.Scrollbar(table_left, orient=tk.HORIZONTAL, command=tree_old_def.xview, style="ColorOld.Horizontal.TScrollbar")
        tree_old_def.configure(yscrollcommand=vsb_old_def.set, xscrollcommand=hsb_old_def.set)
        tree_old_def.grid(row=0, column=0, sticky="nsew")
        vsb_old_def.grid(row=0, column=1, sticky="ns")
        hsb_old_def.grid(row=1, column=0, sticky="ew")
        table_left.grid_rowconfigure(0, weight=1)
        table_left.grid_columnconfigure(0, weight=1)

        # ==================== 右侧：属性维护 ====================
        right_frame = ttk.Frame(main_paned)
        main_paned.add(right_frame, weight=1)
        paned_right = ttk.PanedWindow(right_frame, orient=tk.VERTICAL)
        paned_right.pack(fill=tk.BOTH, expand=True)

        # 上部：新数据
        top_right = ttk.Frame(paned_right)
        paned_right.add(top_right, weight=1)
        ttk.Label(top_right, text="▼ 属性维护 - 即将导入的新数据", font=("微软雅黑", 9, "bold")).pack(anchor=tk.W, padx=5, pady=(5,0))
        tree_new_attr = ttk.Treeview(top_right, columns=("col1","col2","col3"), show="headings", height=6, style="ColorNew.Treeview")
        tree_new_attr.heading("col1", text="属性定义")
        tree_new_attr.heading("col2", text="属性代码")
        tree_new_attr.heading("col3", text="属性描述")
        tree_new_attr.column("col1", width=60, anchor=tk.CENTER)
        tree_new_attr.column("col2", width=100, anchor=tk.CENTER)
        tree_new_attr.column("col3", width=150, anchor=tk.CENTER)
        for row in attr_new:
            tree_new_attr.insert("", tk.END, values=row)
        vsb_new_attr = ttk.Scrollbar(top_right, orient=tk.VERTICAL, command=tree_new_attr.yview)
        hsb_new_attr = ttk.Scrollbar(top_right, orient=tk.HORIZONTAL, command=tree_new_attr.xview)
        tree_new_attr.configure(yscrollcommand=vsb_new_attr.set, xscrollcommand=hsb_new_attr.set)
        tree_new_attr.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
        vsb_new_attr.pack(side=tk.RIGHT, fill=tk.Y)
        hsb_new_attr.pack(side=tk.BOTTOM, fill=tk.X)

        # 下部：旧数据（灰色背景）
        bottom_right = tk.Frame(paned_right, bg='#f0f0f0')
        paned_right.add(bottom_right, weight=1)
        tk.Label(bottom_right, text="▼ 目标文件中已有数据（只读）", font=("微软雅黑", 9, "bold"), bg='#f0f0f0').pack(anchor=tk.W, padx=5, pady=(5,0))
        table_right = tk.Frame(bottom_right, bg='#f0f0f0')
        table_right.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        tree_old_attr = ttk.Treeview(table_right, columns=("col1","col2","col3"), show="headings", height=6, style="ColorOld.Treeview")
        tree_old_attr.heading("col1", text="属性定义")
        tree_old_attr.heading("col2", text="属性代码")
        tree_old_attr.heading("col3", text="属性描述")
        tree_old_attr.column("col1", width=60, anchor=tk.CENTER)
        tree_old_attr.column("col2", width=100, anchor=tk.CENTER)
        tree_old_attr.column("col3", width=150, anchor=tk.CENTER)
        for row in attr_old:
            tree_old_attr.insert("", tk.END, values=row)
        if not attr_old:
            tree_old_attr.insert("", tk.END, values=["(无已有数据)", "", ""])
        vsb_old_attr = ttk.Scrollbar(table_right, orient=tk.VERTICAL, command=tree_old_attr.yview, style="ColorOld.Vertical.TScrollbar")
        hsb_old_attr = ttk.Scrollbar(table_right, orient=tk.HORIZONTAL, command=tree_old_attr.xview, style="ColorOld.Horizontal.TScrollbar")
        tree_old_attr.configure(yscrollcommand=vsb_old_attr.set, xscrollcommand=hsb_old_attr.set)
        tree_old_attr.grid(row=0, column=0, sticky="nsew")
        vsb_old_attr.grid(row=0, column=1, sticky="ns")
        hsb_old_attr.grid(row=1, column=0, sticky="ew")
        table_right.grid_rowconfigure(0, weight=1)
        table_right.grid_columnconfigure(0, weight=1)

        # ---------- 底部按钮 ----------
        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(fill=tk.X, padx=5, pady=10)

        def final_import():
            dialog.destroy()
            self._final_import_colors(color_items)

        ttk.Button(btn_frame, text="最终确认导入", command=final_import).pack(side=tk.LEFT, padx=20)
        ttk.Button(btn_frame, text="取消", command=dialog.destroy).pack(side=tk.LEFT, padx=20)

    def _final_import_colors(self, color_items):
        """实际执行颜色数据写入（保留原有写入逻辑及历史记录）"""
        source_dir = os.path.dirname(self.source_file_path.get())
        color_def_file = os.path.join(source_dir, "颜色定义_批量导入.xlsx")
        attr_file = os.path.join(source_dir, "属性维护_批量导入.xlsx")

        # 记录历史
        history_id = self.db.record_history(
            source_file=self.source_file_path.get(),
            source_sheet='CO',
            total_rows=len(color_items),
            target_files_list=["颜色定义_批量导入.xlsx", "属性维护_批量导入.xlsx"],
            import_type='color'
        )

        # 导入颜色定义
        self._import_color_definition(color_def_file, color_items)
        # 导入属性维护
        self._import_attribute_maintain(attr_file, color_items)

        # 记录明细
        for idx, (code, name) in enumerate(color_items, start=1):
            self.db.record_detail(
                history_id=history_id,
                import_type='color',
                target_file="颜色定义_批量导入.xlsx",
                source_row=idx,
                product_name=name,
                style=code,
                color='',
                status='成功',
                message='',
                row_data_json=json.dumps({'颜色代码': code, '颜色名称': name}, ensure_ascii=False)
            )
        self.log(f"颜色导入完成，共处理 {len(color_items)} 条记录")
        messagebox.showinfo("导入完成", f"已成功导入 {len(color_items)} 条颜色数据到两个模板文件。")

    def _do_import_colors(self, items):
        source_dir = os.path.dirname(self.source_file_path.get())
        color_def_file = os.path.join(source_dir, "颜色定义_批量导入.xlsx")
        attr_file = os.path.join(source_dir, "属性维护_批量导入.xlsx")
        
        # 记录历史
        history_id = self.db.record_history(
            source_file=self.source_file_path.get(),
            source_sheet='CO',
            total_rows=len(items),
            target_files_list=[os.path.basename(color_def_file), os.path.basename(attr_file)],
            import_type='color'
        )
        
        self._import_color_definition(color_def_file, items)
        self._import_attribute_maintain(attr_file, items)
        
        # 记录明细（可选，每条颜色作为一行明细）
        for idx, (code, name) in enumerate(items, start=1):
            self.db.record_detail(
                history_id=history_id,
                import_type='color',
                target_file=os.path.basename(color_def_file),  # 简化只记录一个文件
                source_row=idx,
                product_name=name,
                style=code,
                color='',
                status='成功',
                message='',
                row_data_json=json.dumps({'颜色代码': code, '颜色名称': name}, ensure_ascii=False)
            )
        self.log(f"颜色导入完成，共处理 {len(items)} 条记录")

    def _import_color_definition(self, file_path, items):
        if not os.path.exists(file_path):
            ans = messagebox.askyesno("文件不存在", f"找不到文件：{os.path.basename(file_path)}\n是否手动选择该文件？")
            if ans:
                file_path = filedialog.askopenfilename(
                    title="请选择颜色定义_批量导入.xlsx",
                    initialdir=os.path.dirname(file_path),
                    filetypes=[("Excel files", "*.xlsx")]
                )
                if not file_path:
                    self.log("跳过颜色定义导入")
                    return
            else:
                self.log("跳过颜色定义导入")
                return
        try:
            wb = load_workbook(file_path)
            ws = wb.active
            if ws.max_row < 1 or ws.cell(1, 1).value is None:
                ws.cell(1, 1, "颜色")
                ws.cell(1, 2, "颜色名称")
                ws.cell(1, 3, "备注")
            if ws.max_row >= 2:
                ws.delete_rows(2, ws.max_row - 1)
            for idx, (code, name) in enumerate(items, start=2):
                ws.cell(idx, 1, code).number_format = '@'
                ws.cell(idx, 2, name).number_format = '@'
            wb.save(file_path)
            wb.close()
            self.log(f"颜色定义_批量导入写入 {len(items)} 条记录")
        except Exception as e:
            self.log(f"颜色定义导入失败：{e}")

    def _import_attribute_maintain(self, file_path, items):
        if not os.path.exists(file_path):
            ans = messagebox.askyesno("文件不存在", f"找不到文件：{os.path.basename(file_path)}\n是否手动选择该文件？")
            if ans:
                file_path = filedialog.askopenfilename(
                    title="请选择属性维护_批量导入.xlsx",
                    initialdir=os.path.dirname(file_path),
                    filetypes=[("Excel files", "*.xlsx")]
                )
                if not file_path:
                    self.log("跳过属性维护导入")
                    return
            else:
                self.log("跳过属性维护导入")
                return
        try:
            wb = load_workbook(file_path)
            ws = wb.active
            if ws.max_row < 1 or ws.cell(1, 1).value is None:
                ws.cell(1, 1, "属性定义")
                ws.cell(1, 2, "属性代码")
                ws.cell(1, 3, "属性描述")
            if ws.max_row >= 2:
                ws.delete_rows(2, ws.max_row - 1)
            for idx, (code, name) in enumerate(items, start=2):
                ws.cell(idx, 1, "(颜色)").number_format = '@'
                ws.cell(idx, 2, f"Y{code}").number_format = '@'
                ws.cell(idx, 3, name).number_format = '@'
            wb.save(file_path)
            wb.close()
            self.log(f"属性维护_批量导入模板表写入 {len(items)} 条记录")
        except Exception as e:
            self.log(f"属性维护导入失败：{e}")

    def _prompt_missing_colors(self, missing_colors):
        max_int = 0
        width = 3
        for code in self.color_map.values():
            try:
                val = int(code)
                if val > max_int:
                    max_int = val
                    width = len(code)
            except:
                continue
        code_lines = []
        current = max_int + 1
        for name in missing_colors:
            code = str(current).zfill(width)
            code_lines.append(f"{code}\t{name}")
            current += 1
        content = "\n".join(code_lines)

        top = tk.Toplevel(self.root)
        top.title("颜色代码生成（需手动添加）")
        top.geometry("300x400")
        top.transient(self.root)
        top.grab_set()
        ttk.Label(top, text="以下颜色在CO表中不存在，建议代码已生成。\n请复制颜色和代码，手动粘贴到CO表，\n保存源文件后请在本软件上执行刷新。").pack(pady=5)
        btn_frame = ttk.Frame(top)
        btn_frame.pack(fill=tk.X, pady=5)
        text_frame = ttk.Frame(top)
        text_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        text = tk.Text(text_frame, wrap=tk.NONE, font=("微软雅黑", 10))
        v_scroll = ttk.Scrollbar(text_frame, orient=tk.VERTICAL, command=text.yview)
        h_scroll = ttk.Scrollbar(text_frame, orient=tk.HORIZONTAL, command=text.xview)
        text.configure(yscrollcommand=v_scroll.set, xscrollcommand=h_scroll.set)
        text.grid(row=0, column=0, sticky="nsew")
        v_scroll.grid(row=0, column=1, sticky="ns")
        h_scroll.grid(row=1, column=0, sticky="ew")
        text_frame.grid_rowconfigure(0, weight=1)
        text_frame.grid_columnconfigure(0, weight=1)
        text.insert("1.0", content)

        def select_all():
            text.tag_add(tk.SEL, "1.0", tk.END)
            text.mark_set(tk.INSERT, "1.0")
            text.see(tk.INSERT)
            text.focus_set()
        def copy_selected():
            try:
                selected = text.get(tk.SEL_FIRST, tk.SEL_LAST)
            except:
                messagebox.showwarning("未选中", "请先选中要复制的内容", parent=top)
                return
            if selected.strip():
                top.clipboard_clear()
                top.clipboard_append(selected)
                messagebox.showinfo("复制成功", "选中的内容已复制到剪贴板", parent=top)
            else:
                messagebox.showwarning("空内容", "选中的内容为空", parent=top)
        def cancel():
            top.destroy()
        ttk.Button(btn_frame, text="全选", command=select_all).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="复制", command=copy_selected).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="退出", command=cancel).pack(side=tk.RIGHT, padx=5)
        top.wait_window()
        return None

    def show_supplier_import_dialog(self):
        """显示供应商导入对话框，预览并确认导入（显示最后20条，支持多选）"""
        if not self.source_loader:
            messagebox.showerror("错误", "请先加载源文件")
            return
        if '供应商' not in self.source_loader.wb.sheetnames:
            messagebox.showerror("错误", "源文件中未找到'供应商'工作表")
            return

        ws = self.source_loader.wb['供应商']
        # 获取表头（第一行）
        headers = []
        for col in range(1, ws.max_column + 1):
            val = ws.cell(1, col).value
            headers.append(str(val).strip() if val else f"列{col}")
        
        # 读取所有有效数据行（从第2行开始）
        all_data_rows = []
        for row in range(2, ws.max_row + 1):
            row_data = []
            for col in range(1, ws.max_column + 1):
                cell_val = ws.cell(row, col).value
                row_data.append(str(cell_val).strip() if cell_val is not None else "")
            if any(row_data):
                all_data_rows.append(row_data)

        if not all_data_rows:
            messagebox.showinfo("提示", "供应商工作表中没有有效的数据")
            return

        # 取最后20条数据用于预览
        preview_rows = all_data_rows[-20:] if len(all_data_rows) > 20 else all_data_rows

        dialog = tk.Toplevel(self.root)
        dialog.title("供应商数据预览 - 最后20条记录（可多选）")
        dialog.geometry("1300x750")
        dialog.transient(self.root)
        dialog.grab_set()

        info_text = (f"共 {len(all_data_rows)} 条供应商数据，此处显示最后 {len(preview_rows)} 条。\n"
                     f"将导入到：\n"
                     f"  • 供应商维护_批量导入-庄子.xlsx（保留源编号）\n"
                     f"  • 供应商维护_批量导入-帛点.xlsx（编号自动递增 GYSxxxx）\n"
                     f"请勾选需要导入的行（可多选）。")
        ttk.Label(dialog, text=info_text, font=("微软雅黑", 10)).pack(pady=10)

        frame = ttk.Frame(dialog)
        frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        columns = [f"col{i}" for i in range(len(headers))]
        tree = ttk.Treeview(frame, columns=columns, show="headings", height=12, selectmode='extended')
        for i, header in enumerate(headers):
            tree.heading(columns[i], text=header)
            tree.column(columns[i], width=100, anchor=tk.CENTER)
        
        # 插入预览数据（同时保存完整原始数据索引）
        tree.preview_indices = []
        start_idx = len(all_data_rows) - len(preview_rows)
        for i, row_data in enumerate(preview_rows):
            tree.insert("", tk.END, values=row_data)
            tree.preview_indices.append(start_idx + i)
        
        # 自动滚动到最末行
        children = tree.get_children()
        if children:
            tree.see(children[-1])

        vsb = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)
        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        frame.grid_rowconfigure(0, weight=1)
        frame.grid_columnconfigure(0, weight=1)

        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(fill=tk.X, padx=5, pady=10)

        def on_tree_select(event):
            if tree.selection():
                import_btn.config(state=tk.NORMAL)
            else:
                import_btn.config(state=tk.DISABLED)

        tree.bind("<<TreeviewSelect>>", on_tree_select)

        def do_import():
            selected_items = tree.selection()
            if not selected_items:
                return
            selected_indices = [tree.preview_indices[tree.index(item)] for item in selected_items]
            selected_rows = [all_data_rows[i] for i in selected_indices]
            dialog.destroy()
            # 改为打开导入结果预览窗口，而非直接写入
            self.show_import_result_preview(headers, selected_rows)

        import_btn = ttk.Button(btn_frame, text="确认", command=do_import, state=tk.DISABLED)
        import_btn.pack(side=tk.LEFT, padx=20)
        ttk.Button(btn_frame, text="取消", command=dialog.destroy).pack(side=tk.LEFT, padx=20)
        
    def show_import_result_preview(self, headers, selected_rows):
        """供应商导入预览：左右并排（帛点/庄子），每个区域内上下布局（新数据在上，旧数据在下），共用水平滚动条"""
        source_dir = os.path.dirname(self.source_file_path.get())
        zhuangzi_file = "供应商维护_批量导入-庄子.xlsx"
        bodian_file = "供应商维护_批量导入-帛点.xlsx"

        # 样式定义（背景区分）
        style = ttk.Style()
        style.configure("New.Treeview", background="white", fieldbackground="white")
        style.configure("Old.Treeview",
                        background="#f0f0f0", fieldbackground="#f0f0f0", foreground="black")
        style.map("Old.Treeview",
                  background=[('selected', '#c0c0c0')],
                  fieldbackground=[('selected', '#c0c0c0')])
        style.configure("Old.Treeview.Heading", background="#dcdcdc", foreground="black")
        style.configure("Old.Vertical.TScrollbar", troughcolor="#f0f0f0", background="#d0d0d0")
        style.configure("Old.Horizontal.TScrollbar", troughcolor="#f0f0f0", background="#d0d0d0")

        # ---------- 读取现有数据 ----------
        def load_existing_data(file_path, expected_headers):
            if not os.path.exists(file_path):
                return [], []
            try:
                wb = load_workbook(file_path)
                ws = wb.active
                existing_headers = []
                for col in range(1, ws.max_column + 1):
                    val = ws.cell(1, col).value
                    existing_headers.append(str(val).strip() if val else f"列{col}")
                header_to_col = {h: idx for idx, h in enumerate(existing_headers)}
                data_rows = []
                for row in range(2, ws.max_row + 1):
                    row_data = []
                    for h in expected_headers:
                        if h in header_to_col:
                            col_idx = header_to_col[h] + 1
                            val = ws.cell(row, col_idx).value
                            row_data.append(str(val).strip() if val is not None else "")
                        else:
                            row_data.append("")
                    if any(row_data):
                        data_rows.append(row_data)
                wb.close()
                return existing_headers, data_rows
            except Exception as e:
                self.log(f"读取现有数据失败 {file_path}: {e}")
                return [], []

        # 计算帛点新编号
        bodian_max_num = 0
        bodian_path = os.path.join(source_dir, bodian_file)
        if os.path.exists(bodian_path):
            try:
                wb = load_workbook(bodian_path)
                ws = wb.active
                if ws and ws.max_row >= 2:
                    for row in range(2, ws.max_row + 1):
                        code_cell = ws.cell(row, 1).value
                        if code_cell and isinstance(code_cell, str) and code_cell.startswith("GYS"):
                            num_part = code_cell[3:]
                            if num_part.isdigit():
                                num = int(num_part)
                                if num > bodian_max_num:
                                    bodian_max_num = num
                wb.close()
            except Exception:
                pass

        preview_data_zhuangzi = []
        preview_data_bodian = []
        for row_data in selected_rows:
            row = list(row_data)
            while len(row) < len(headers):
                row.append("")
            preview_data_zhuangzi.append(row)
            bodian_max_num += 1
            new_code = f"GYS{bodian_max_num:04d}"
            row_bodian = list(row)
            row_bodian[0] = new_code
            preview_data_bodian.append(row_bodian)

        bodian_existing_headers, bodian_existing_data = load_existing_data(bodian_path, headers)
        zhuangzi_existing_headers, zhuangzi_existing_data = load_existing_data(os.path.join(source_dir, zhuangzi_file), headers)

        # ---------- 创建主窗口 ----------
        dialog = tk.Toplevel(self.root)
        dialog.title("供应商导入预览 - 最终确认（双击帛点供应商编号可修改）")
        dialog.geometry("1400x750")
        dialog.transient(self.root)
        dialog.grab_set()

        ttk.Label(dialog, text=f"即将导入 {len(selected_rows)} 条供应商数据，请最终确认：\n"
                               f"  • 供应商维护_批量导入-帛点.xlsx —— 编号自动递增 GYSxxxx（双击可修改）\n"
                               f"  • 供应商维护_批量导入-庄子.xlsx —— 保留源编号（不可编辑）\n"
                               f"左右两侧分别展示两个文件将写入的数据（新旧对比）。",
                  font=("微软雅黑", 10)).pack(pady=10)

        # 左右主分割
        main_paned = ttk.PanedWindow(dialog, orient=tk.HORIZONTAL)
        main_paned.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # ==================== 左侧：帛点 ====================
        left_frame = ttk.Frame(main_paned)
        main_paned.add(left_frame, weight=1)
        paned_left = ttk.PanedWindow(left_frame, orient=tk.VERTICAL)
        paned_left.pack(fill=tk.BOTH, expand=True)

        # ---- 新数据表格 ----
        top_left = ttk.Frame(paned_left)
        paned_left.add(top_left, weight=1)
        ttk.Label(top_left, text="▼ 帛点 - 即将导入的新数据（双击第一列可编辑）", font=("微软雅黑", 9, "bold")).pack(anchor=tk.W, padx=5, pady=(5,0))
        tree_new_b = ttk.Treeview(top_left, columns=[f"col{i}" for i in range(len(headers))], show="headings", height=8, style="New.Treeview")
        for i, header in enumerate(headers):
            tree_new_b.heading(f"col{i}", text=header)
            tree_new_b.column(f"col{i}", width=100, anchor=tk.CENTER)
        for row in preview_data_bodian:
            tree_new_b.insert("", tk.END, values=row)
        vsb_new_b = ttk.Scrollbar(top_left, orient=tk.VERTICAL, command=tree_new_b.yview)
        tree_new_b.configure(yscrollcommand=vsb_new_b.set)
        tree_new_b.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vsb_new_b.pack(side=tk.RIGHT, fill=tk.Y)

        # ---- 旧数据表格 ----
        bottom_left = tk.Frame(paned_left, bg='#f0f0f0')
        paned_left.add(bottom_left, weight=1)
        tk.Label(bottom_left, text="▼ 目标文件中已有数据（只读）", font=("微软雅黑", 9, "bold"), bg='#f0f0f0').pack(anchor=tk.W, padx=5, pady=(5,0))

        # 表格和水平滚动条容器
        table_left = tk.Frame(bottom_left, bg='#f0f0f0')
        table_left.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        tree_old_b = ttk.Treeview(table_left, columns=[f"col{i}" for i in range(len(headers))], show="headings", height=8, style="Old.Treeview")
        for i, header in enumerate(headers):
            tree_old_b.heading(f"col{i}", text=header)
            tree_old_b.column(f"col{i}", width=100, anchor=tk.CENTER)
        for row in bodian_existing_data:
            tree_old_b.insert("", tk.END, values=row)
        if not bodian_existing_data:
            tree_old_b.insert("", tk.END, values=["(无已有数据)"] + [""]*(len(headers)-1))
        vsb_old_b = ttk.Scrollbar(table_left, orient=tk.VERTICAL, command=tree_old_b.yview, style="Old.Vertical.TScrollbar")
        hsb_b = ttk.Scrollbar(table_left, orient=tk.HORIZONTAL, command=lambda *args: (tree_new_b.xview(*args), tree_old_b.xview(*args)))
        tree_new_b.configure(xscrollcommand=hsb_b.set)
        tree_old_b.configure(xscrollcommand=hsb_b.set)
        tree_old_b.grid(row=0, column=0, sticky="nsew")
        vsb_old_b.grid(row=0, column=1, sticky="ns")
        hsb_b.grid(row=1, column=0, sticky="ew")
        table_left.grid_rowconfigure(0, weight=1)
        table_left.grid_columnconfigure(0, weight=1)

        # 帛点新数据表格的双击编辑功能
        def on_b_double_click(event):
            region = tree_new_b.identify_region(event.x, event.y)
            if region != "cell":
                return
            column = tree_new_b.identify_column(event.x)
            if column != "#1":
                return
            item = tree_new_b.selection()[0] if tree_new_b.selection() else tree_new_b.identify_row(event.y)
            if not item:
                return
            col_idx = 0
            bbox = tree_new_b.bbox(item, column)
            if not bbox:
                return
            x, y, width, height = bbox
            current_value = tree_new_b.item(item, "values")[col_idx]

            entry = ttk.Entry(tree_new_b)
            entry.place(x=x, y=y, width=width, height=height)
            entry.insert(0, current_value)
            entry.select_range(0, tk.END)
            entry.focus()

            def on_confirm(event=None):
                new_val = entry.get().strip()
                if new_val:
                    values = list(tree_new_b.item(item, "values"))
                    values[col_idx] = new_val
                    tree_new_b.item(item, values=values)
                    idx = tree_new_b.index(item)
                    preview_data_bodian[idx][0] = new_val
                entry.destroy()

            def on_cancel(event=None):
                entry.destroy()

            entry.bind("<Return>", on_confirm)
            entry.bind("<Escape>", on_cancel)
            entry.bind("<FocusOut>", lambda e: entry.destroy())

        tree_new_b.bind("<Double-1>", on_b_double_click)

        # ==================== 右侧：庄子 ====================
        right_frame = ttk.Frame(main_paned)
        main_paned.add(right_frame, weight=1)
        paned_right = ttk.PanedWindow(right_frame, orient=tk.VERTICAL)
        paned_right.pack(fill=tk.BOTH, expand=True)

        # 新数据
        top_right = ttk.Frame(paned_right)
        paned_right.add(top_right, weight=1)
        ttk.Label(top_right, text="▼ 庄子 - 即将导入的新数据（只读）", font=("微软雅黑", 9, "bold")).pack(anchor=tk.W, padx=5, pady=(5,0))
        tree_new_z = ttk.Treeview(top_right, columns=[f"col{i}" for i in range(len(headers))], show="headings", height=8, style="New.Treeview")
        for i, header in enumerate(headers):
            tree_new_z.heading(f"col{i}", text=header)
            tree_new_z.column(f"col{i}", width=100, anchor=tk.CENTER)
        for row in preview_data_zhuangzi:
            tree_new_z.insert("", tk.END, values=row)
        vsb_new_z = ttk.Scrollbar(top_right, orient=tk.VERTICAL, command=tree_new_z.yview)
        tree_new_z.configure(yscrollcommand=vsb_new_z.set)
        tree_new_z.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vsb_new_z.pack(side=tk.RIGHT, fill=tk.Y)

        # 旧数据
        bottom_right = tk.Frame(paned_right, bg='#f0f0f0')
        paned_right.add(bottom_right, weight=1)
        tk.Label(bottom_right, text="▼ 目标文件中已有数据（只读）", font=("微软雅黑", 9, "bold"), bg='#f0f0f0').pack(anchor=tk.W, padx=5, pady=(5,0))

        table_right = tk.Frame(bottom_right, bg='#f0f0f0')
        table_right.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        tree_old_z = ttk.Treeview(table_right, columns=[f"col{i}" for i in range(len(headers))], show="headings", height=8, style="Old.Treeview")
        for i, header in enumerate(headers):
            tree_old_z.heading(f"col{i}", text=header)
            tree_old_z.column(f"col{i}", width=100, anchor=tk.CENTER)
        for row in zhuangzi_existing_data:
            tree_old_z.insert("", tk.END, values=row)
        if not zhuangzi_existing_data:
            tree_old_z.insert("", tk.END, values=["(无已有数据)"] + [""]*(len(headers)-1))
        vsb_old_z = ttk.Scrollbar(table_right, orient=tk.VERTICAL, command=tree_old_z.yview, style="Old.Vertical.TScrollbar")
        hsb_z = ttk.Scrollbar(table_right, orient=tk.HORIZONTAL, command=lambda *args: (tree_new_z.xview(*args), tree_old_z.xview(*args)))
        tree_new_z.configure(xscrollcommand=hsb_z.set)
        tree_old_z.configure(xscrollcommand=hsb_z.set)
        tree_old_z.grid(row=0, column=0, sticky="nsew")
        vsb_old_z.grid(row=0, column=1, sticky="ns")
        hsb_z.grid(row=1, column=0, sticky="ew")
        table_right.grid_rowconfigure(0, weight=1)
        table_right.grid_columnconfigure(0, weight=1)

        # 底部按钮
        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(fill=tk.X, padx=5, pady=10)

        def final_import():
            dialog.destroy()
            self._do_import_suppliers_final(headers, selected_rows, preview_data_zhuangzi, preview_data_bodian)

        ttk.Button(btn_frame, text="最终确认导入", command=final_import).pack(side=tk.LEFT, padx=20)
        ttk.Button(btn_frame, text="取消", command=dialog.destroy).pack(side=tk.LEFT, padx=20)
        
    def _do_import_suppliers_final(self, headers, selected_rows, data_zhuangzi, data_bodian):
        """实际执行写入操作（使用预览时已准备好的数据）"""
        source_dir = os.path.dirname(self.source_file_path.get())
        target_files = [
            ("供应商维护_批量导入-庄子.xlsx", data_zhuangzi),
            ("供应商维护_批量导入-帛点.xlsx", data_bodian)
        ]
        success_count = 0
        
        # 记录历史
        history_id = self.db.record_history(
            source_file=self.source_file_path.get(),
            source_sheet='供应商',
            total_rows=len(selected_rows),
            target_files_list=[fname for fname, _ in target_files],
            import_type='supplier'
        )

        for fname, data_rows in target_files:
            file_path = os.path.join(source_dir, fname)
            if not os.path.exists(file_path):
                ans = messagebox.askyesno("文件不存在",
                                          f"在源文件目录找不到：{fname}\n是否手动选择该文件？")
                if ans:
                    new_path = filedialog.askopenfilename(
                        title=f"请选择{fname}",
                        initialdir=source_dir,
                        filetypes=[("Excel files", "*.xlsx")]
                    )
                    if new_path:
                        file_path = new_path
                    else:
                        self.log(f"跳过：{fname}")
                        continue
                else:
                    self.log(f"跳过：{fname}")
                    continue

            try:
                wb = load_workbook(file_path)
                ws = wb.active
                if not ws:
                    ws = wb.create_sheet("Sheet1", 0)
                    ws.title = "Sheet1"
                
                # 清空现有数据
                if ws.max_row > 1:
                    ws.delete_rows(2, ws.max_row - 1)
                
                # 写入表头
                for col_idx, header in enumerate(headers, start=1):
                    ws.cell(1, col_idx, header)
                
                # 写入数据行
                for i, row_data in enumerate(data_rows):
                    for col_idx, value in enumerate(row_data[:len(headers)], start=1):
                        cell = ws.cell(i + 2, col_idx, value)
                        cell.number_format = '@'
                
                wb.save(file_path)
                wb.close()
                self.log(f"供应商数据已成功写入：{fname}")
                success_count += 1
                
                # 记录明细（可选，每条供应商一行）
                for idx, row_data in enumerate(data_rows, start=1):
                    supplier_code = row_data[0] if len(row_data) > 0 else ''
                    supplier_name = row_data[1] if len(row_data) > 1 else ''
                    self.db.record_detail(
                        history_id=history_id,
                        import_type='supplier',
                        target_file=fname,
                        source_row=idx,
                        product_name=supplier_name,
                        style=supplier_code,
                        color='',
                        status='成功',
                        message='',
                        row_data_json=json.dumps(dict(zip(headers, row_data)), ensure_ascii=False)
                    )
            except Exception as e:
                self.log(f"写入 {fname} 失败：{e}")
                messagebox.showerror("导入错误", f"写入 {fname} 时出错：{e}")

        self.log(f"供应商导入完成，成功写入 {success_count} 个文件。")
        if success_count == 2:
            messagebox.showinfo("导入完成", "选中的供应商数据已成功导入到两个模板文件。")

    # -------------------- 导入历史 --------------------
    def show_import_history(self):
        if not os.path.exists(self.db.db_path):
            messagebox.showinfo("提示", "暂无导入记录")
            return
        history_win = tk.Toplevel(self.root)
        history_win.title("导入历史记录")
        history_win.geometry("1000x600")
        history_win.transient(self.root)

        filter_frame = ttk.Frame(history_win)
        filter_frame.pack(fill=tk.X, padx=5, pady=5)
        ttk.Label(filter_frame, text="日期筛选:").pack(side=tk.LEFT, padx=5)
        start_date_var = tk.StringVar()
        end_date_var = tk.StringVar()
        ttk.Entry(filter_frame, textvariable=start_date_var, width=12).pack(side=tk.LEFT, padx=2)
        ttk.Label(filter_frame, text="至").pack(side=tk.LEFT)
        ttk.Entry(filter_frame, textvariable=end_date_var, width=12).pack(side=tk.LEFT, padx=2)
        ttk.Label(filter_frame, text="目标文件:").pack(side=tk.LEFT, padx=5)
        file_filter_var = tk.StringVar()
        files = self.db.get_distinct_target_files()
        file_combo = ttk.Combobox(filter_frame, textvariable=file_filter_var, width=20, state="readonly")
        file_combo['values'] = ['全部'] + files
        file_combo.set('全部')
        file_combo.pack(side=tk.LEFT, padx=2)
        ttk.Label(filter_frame, text="导入类型:").pack(side=tk.LEFT, padx=5)
        import_type_var = tk.StringVar(value="全部")
        type_combo = ttk.Combobox(filter_frame, textvariable=import_type_var, values=["全部", "数据导入", "颜色导入", "供应商导入"], state="readonly", width=12)
        type_combo.pack(side=tk.LEFT, padx=2)

        list_frame = ttk.Frame(history_win)
        list_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        scrollbar = ttk.Scrollbar(list_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        history_listbox = tk.Listbox(list_frame, yscrollcommand=scrollbar.set, selectmode=tk.EXTENDED)
        history_listbox._data = []
        history_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=history_listbox.yview)

        btn_frame = ttk.Frame(history_win)
        btn_frame.pack(fill=tk.X, padx=5, pady=5)

        def refresh_history():
            filters = {}
            if start_date_var.get():
                filters['start_date'] = start_date_var.get()
            if end_date_var.get():
                filters['end_date'] = end_date_var.get()
            if file_filter_var.get() and file_filter_var.get() != '全部':
                filters['target_file'] = file_filter_var.get()
            selected_type = import_type_var.get()
            if selected_type == "数据导入":
                filters['import_type'] = 'data'
            elif selected_type == "颜色导入":
                filters['import_type'] = 'color'
            elif selected_type == "供应商导入":
                filters['import_type'] = 'supplier'
            rows = self.db.get_history(filters)
            history_listbox.delete(0, tk.END)
            history_listbox._data = []
            for row in rows:
                # row: id, import_time, source_file, source_sheet, total_rows, target_files, import_type
                type_display = {'data':'数据', 'color':'颜色', 'supplier':'供应商'}.get(row[6], row[6])
                display = f"[{type_display}] {row[1]} | {row[2]} | {row[3]} | 共{row[4]}行 | 目标:{row[5]}"
                history_listbox.insert(tk.END, display)
                history_listbox._data.append(row[0])

        def clean_records():
            selected = history_listbox.curselection()
            if not selected:
                messagebox.showwarning("提示", "请先选择要删除的历史记录", parent=history_win)
                return
            ids_to_delete = [history_listbox._data[i] for i in selected]
            pwd = simpledialog.askstring("密码验证", "请输入清理密码:", parent=history_win, show='*')
            if pwd != "zg1018":
                messagebox.showerror("错误", "密码错误，无法清理", parent=history_win)
                return
            if messagebox.askyesno("确认删除", f"确定要删除 {len(ids_to_delete)} 条历史记录吗？\n此操作不可恢复！", parent=history_win):
                for hid in ids_to_delete:
                    self.db.delete_history(hid)
                self.db.vacuum()
                refresh_history()
                messagebox.showinfo("完成", f"已删除 {len(ids_to_delete)} 条记录，数据库空间已回收", parent=history_win)

        ttk.Button(btn_frame, text="查询", command=refresh_history).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="查看明细", command=lambda: self.show_detail(history_win, history_listbox)).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="记录清理", command=clean_records).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="关闭", command=history_win.destroy).pack(side=tk.RIGHT, padx=5)

        refresh_history()

    def show_detail(self, parent_win, history_listbox):
        selection = history_listbox.curselection()
        if not selection:
            messagebox.showwarning("提示", "请先选择一条历史记录", parent=parent_win)
            return
        history_id = history_listbox._data[selection[0]]
        detail_win = tk.Toplevel(parent_win)
        detail_win.title(f"导入明细 - 记录ID {history_id}")
        detail_win.geometry("1100x700")
        detail_win.transient(parent_win)

        all_details = self.db.get_details(history_id)
        target_files = sorted(set(det[1] for det in all_details))
        
        detail_win.current_filter = None
        detail_win.current_find_idx = -1

        top_frame = ttk.Frame(detail_win)
        top_frame.pack(fill=tk.X, padx=5, pady=5)

        ttk.Label(top_frame, text="目标文件：").pack(side=tk.LEFT, padx=(0,5))
        filter_var = tk.StringVar(value="全部")
        filter_combo = ttk.Combobox(top_frame, textvariable=filter_var, values=["全部"] + target_files, state="readonly", width=25)
        filter_combo.pack(side=tk.LEFT, padx=5)

        ttk.Label(top_frame, text="查找：").pack(side=tk.LEFT, padx=(15,5))
        find_entry = ttk.Entry(top_frame, width=20)
        find_entry.pack(side=tk.LEFT, padx=5)
        find_next_btn = ttk.Button(top_frame, text="下一个")
        find_next_btn.pack(side=tk.LEFT, padx=2)
        find_prev_btn = ttk.Button(top_frame, text="上一个")
        find_prev_btn.pack(side=tk.LEFT, padx=2)

        paned = ttk.PanedWindow(detail_win, orient=tk.HORIZONTAL)
        paned.pack(fill=tk.BOTH, expand=True, pady=(0,15))

        left_frame = ttk.Frame(paned)
        paned.add(left_frame, weight=1)
        columns = ("source_row", "target_file", "product_name", "style", "color", "status", "message")
        tree = ttk.Treeview(left_frame, columns=columns, show="headings")
        for col in columns:
            tree.heading(col, text=col)
            if col == "source_row":
                tree.column(col, width=50)
            elif col == "target_file":
                tree.column(col, width=160)
            elif col == "product_name":
                tree.column(col, width=110)
            elif col == "style":
                tree.column(col, width=150)
            elif col == "color":
                tree.column(col, width=120)
            elif col == "status":
                tree.column(col, width=60)
            else:
                tree.column(col, width=80)
        vsb = ttk.Scrollbar(left_frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)

        right_frame = ttk.Frame(paned)
        paned.add(right_frame, weight=1)
        ttk.Label(right_frame, text="源文件完整数据（选中明细行后自动显示）", font=("微软雅黑", 10, "bold")).pack(anchor=tk.W, pady=5)
        data_text = tk.Text(right_frame, wrap=tk.WORD, font=("微软雅黑", 9))
        data_text.pack(fill=tk.BOTH, expand=True)
        data_scroll = ttk.Scrollbar(right_frame, orient=tk.VERTICAL, command=data_text.yview)
        data_text.config(yscrollcommand=data_scroll.set)
        data_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        def refresh_treeview():
            for item in tree.get_children():
                tree.delete(item)
            if detail_win.current_filter is None or detail_win.current_filter == "全部":
                data_source = all_details
            else:
                data_source = [det for det in all_details if det[1] == detail_win.current_filter]
            for det in data_source:
                tree.insert("", tk.END, values=det[:7])
            detail_win.current_find_idx = -1

        def on_filter_change(event=None):
            selected = filter_var.get()
            if selected == "全部":
                detail_win.current_filter = None
            else:
                detail_win.current_filter = selected
            refresh_treeview()

        def find_text(direction=1):
            keyword = find_entry.get().strip()
            if not keyword:
                messagebox.showinfo("提示", "请输入要查找的内容", parent=detail_win)
                return
            items = tree.get_children()
            if not items:
                return
            matches = []
            for idx, item in enumerate(items):
                values = tree.item(item, "values")
                full_text = " ".join(str(v) for v in values).lower()
                if keyword.lower() in full_text:
                    matches.append(idx)
            if not matches:
                messagebox.showinfo("提示", f"未找到包含“{keyword}”的记录", parent=detail_win)
                detail_win.current_find_idx = -1
                return
            current_sel = tree.selection()
            current_pos = -1
            if current_sel:
                try:
                    current_pos = items.index(current_sel[0])
                except ValueError:
                    current_pos = -1
            if direction == 1:
                for pos in matches:
                    if pos > current_pos:
                        target_pos = pos
                        break
                else:
                    target_pos = matches[0]
            else:
                for pos in reversed(matches):
                    if pos < current_pos:
                        target_pos = pos
                        break
                else:
                    target_pos = matches[-1]
            target_item = items[target_pos]
            tree.selection_set(target_item)
            tree.see(target_item)
            tree.event_generate("<<TreeviewSelect>>")
            detail_win.current_find_idx = target_pos

        filter_combo.bind("<<ComboboxSelected>>", on_filter_change)
        find_next_btn.config(command=lambda: find_text(1))
        find_prev_btn.config(command=lambda: find_text(-1))
        find_entry.bind("<Return>", lambda e: find_text(1))

        refresh_treeview()

        def on_tree_select(event):
            selected = tree.selection()
            if not selected:
                return
            item = selected[0]
            values = tree.item(item, "values")
            source_row = values[0]
            target_file = values[1]
            with sqlite3.connect(self.db.db_path) as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT source_row_data FROM import_details WHERE history_id=? AND source_row=? AND target_file=?", (history_id, source_row, target_file))
                res = cursor.fetchone()
            if res and res[0]:
                try:
                    data_dict = json.loads(res[0])
                    display = "\n".join(f"{k}: {v}" for k, v in data_dict.items())
                    data_text.delete(1.0, tk.END)
                    data_text.insert(tk.END, display)
                except:
                    data_text.delete(1.0, tk.END)
                    data_text.insert(tk.END, "数据解析失败")
            else:
                data_text.delete(1.0, tk.END)
                data_text.insert(tk.END, "无源行数据")

        tree.bind("<<TreeviewSelect>>", on_tree_select)

        def adjust_pane_ratio(event=None):
            width = paned.winfo_width()
            if width <= 0:
                return
            target_pos = int(width * 7 / 10)
            current_pos = paned.sashpos(0)
            if abs(current_pos - target_pos) > 2:
                paned.sashpos(0, target_pos)

        paned.bind("<Configure>", adjust_pane_ratio)
        detail_win.after(10, adjust_pane_ratio)
    
    # -------------------- 日志 --------------------
    def log(self, message):
        def _log():
            self.log_text.config(state=tk.NORMAL)
            self.log_text.insert(tk.END, message + "\n")
            self.log_text.see(tk.END)
            self.log_text.config(state=tk.DISABLED)
        self.root.after(0, _log)

    def clear_log(self):
        if messagebox.askyesno("确认", "确定要清空所有日志信息吗？"):
            self._clear_log_content()
            self.log("日志已清空")

    def _clear_log_content(self):
        self.log_text.config(state=tk.NORMAL)
        self.log_text.delete(1.0, tk.END)
        self.log_text.config(state=tk.DISABLED)

    def show_help(self):
        help_msg = """
=====数据检查功能=====
    1、成分含量检查：检查成分百分比总和是否为100的整数倍，单个成分含量是否≤100%，含量合计值>500%时预警提示
    2、成分违规词检查：检查面料/里料/其他成分中是否含有预设的违规词汇
    3、执行标准检查：检查标准数据和预设值是否相符
    4、价格检查：标准价是否大于采购价、倍率范围超出预警值(2-20)提示、采购价标准价超出预警值提示、价格小数位数是否合规，防止价格录入错误
    5、重复检查：同款号的价格、执行标准等一致性检查
    6、格式检查：检查内容是否含空格、空白字符、英文冒号，自动删除空格、替换英文冒号为中文冒号
    7、商品级别检查：大类为皮装时商品级别是否为皮装

=====数据导入功能=====
    1、商品信息自动导入：将选中的商品信息数据自动导入亿博士商品信息模板表
    2、条码数据自动导入：选中的款号和自动匹配的颜色代码导入亿博士条码批量生成模板表
    3、将选中的颜色代码和名称导入模板表：“颜色定义_批量导入.xlsx”和“属性维护_批量导入.xlsx”

=====注意事项=====
    1、执行标准源表数据更新时间2026年2月，如有新发布执行标准，源表数据未及时更新，注意以官方新发布的为准！
    2、使用本软件过程中若源文件有更新，需及时在本软件点击更新按钮执行数据更新，防止数据错误。

                                                                                    祝钢  2026年3月10日
"""
        messagebox.showinfo("功能说明", help_msg)


# ==================== 启动与授权 ====================
def check_password(input_pw, current_date):
    return input_pw == f"zg{current_date.year}#"

def main():
    current_date = datetime.now().date()
    FREE_UNTIL = date(2026, 12, 31)

    root = tk.Tk()
    root.withdraw()

    if current_date <= FREE_UNTIL:
        root.deiconify()
        app = DataCheckerImporter(root)
        root.mainloop()
        return

    license_file = os.path.join(os.path.dirname(sys.argv[0]), "license.dat")
    
    # 加载或初始化授权数据
    data = {}
    if os.path.exists(license_file):
        try:
            with open(license_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
        except:
            pass
    
    # 补全缺失字段
    data.setdefault('auth_date', None)
    data.setdefault('last_run', None)
    data.setdefault('total_errors', 0)
    data.setdefault('lock_until', None)   # 锁定截止时间戳（浮点字符串）
    
    # 检查锁定状态
    lock_until = data.get('lock_until')
    if lock_until:
        try:
            lock_ts = float(lock_until)
            if time.time() < lock_ts:
                remain = lock_ts - time.time()
                if remain < 3600:
                    msg = f"因多次密码错误，程序已锁定，请等待 {int(remain//60)} 分 {int(remain%60)} 秒后重试。"
                else:
                    hours = int(remain // 3600)
                    minutes = int((remain % 3600) // 60)
                    msg = f"因多次密码错误，程序已锁定，请等待 {hours} 小时 {minutes} 分后重试。"
                messagebox.showerror("锁定", msg, parent=root)
                root.destroy()
                sys.exit(0)
        except:
            pass
    
    # 检查现有授权是否有效（一年内）
    valid = False
    if data.get('auth_date') and data.get('last_run'):
        try:
            auth_date = datetime.strptime(data['auth_date'], "%Y-%m-%d").date()
            last_run = datetime.strptime(data['last_run'], "%Y-%m-%d").date()
            if current_date < last_run:
                raise ValueError("TIMEYIBEIXIUGAI")
            if (current_date - auth_date).days <= 365:
                valid = True
        except:
            pass
    
    if not valid:
        # 需要输入密码
        password = simpledialog.askstring("验证", f"当前日期：{current_date}\n请输入密码:", parent=root, show='*')
        if password is None:
            root.destroy()
            sys.exit(0)
        
        if not check_password(password, current_date):
            # 密码错误，增加错误计数
            data['total_errors'] = data.get('total_errors', 0) + 1
            errors = data['total_errors']
            lock_until = None
            if errors >= 10:
                lock_until = time.time() + 86400  # 24小时
                msg = "密码错误已达10次，程序已锁定24小时。"
            elif errors >= 3:
                lock_until = time.time() + 180   # 3分钟
                msg = "密码错误已达3次，程序已锁定3分钟。"
            else:
                msg = f"密码错误，还剩 {3 - errors} 次机会触发3分钟锁，{10 - errors} 次机会触发24小时锁。"
            
            if lock_until:
                data['lock_until'] = str(lock_until)
                # 注意：不重置错误计数，解锁后继续累积，以便触发更高阈值
            # 保存更新后的数据（保留原有 auth_date 等字段）
            with open(license_file, 'w', encoding='utf-8') as f:
                json.dump({
                    'auth_date': data.get('auth_date'),
                    'last_run': data.get('last_run'),
                    'total_errors': data['total_errors'],
                    'lock_until': data.get('lock_until')
                }, f)
            messagebox.showerror("错误", f"{msg}\n程序将退出。", parent=root)
            root.destroy()
            sys.exit(1)
        else:
            # 密码正确，重置错误计数和锁定
            data['total_errors'] = 0
            data['lock_until'] = None
            # 如果从未授权，则设置授权起始日期为今天
            if not data.get('auth_date'):
                data['auth_date'] = current_date.strftime("%Y-%m-%d")
            data['last_run'] = current_date.strftime("%Y-%m-%d")
            with open(license_file, 'w', encoding='utf-8') as f:
                json.dump(data, f)
    else:
        # 已有有效授权，仅更新最后运行日期，同时清除残留锁定信息
        data['last_run'] = current_date.strftime("%Y-%m-%d")
        data['total_errors'] = 0
        data['lock_until'] = None
        with open(license_file, 'w', encoding='utf-8') as f:
            json.dump(data, f)
    
    root.deiconify()
    app = DataCheckerImporter(root)
    root.mainloop()

if __name__ == "__main__":
    main()