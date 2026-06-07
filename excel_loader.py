import re
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

class ColumnMapper:
    def __init__(self, ws):
        self.ws = ws
        self.headers = {}
        self.col_index = {}
        self._build_headers()

    def _build_headers(self):
        for col in range(1, self.ws.max_column + 1):
            val = self.ws.cell(1, col).value
            self.headers[col] = str(val).strip() if val else f"列{col}"

    def map_columns(self, keyword_dict):
        for std_key, keywords in keyword_dict.items():
            for col, header in self.headers.items():
                for kw in keywords:
                    if kw in header:
                        self.col_index[std_key] = col
                        break
                if std_key in self.col_index:
                    break

    def get_col(self, std_key):
        return self.col_index.get(std_key)


class ExcelSourceLoader:
    def __init__(self, file_path):
        self.file_path = file_path
        self.wb = None
        self.ws = None
        self.col_mapper = None
        self.cached_data = {}
        self.allowed_sheets = ["帛点", "庄子", "样衣"]
        self.keyword_map = {
            '款号': ['款号'],
            '品名': ['品名'],
            '颜色': ['颜色', '(颜色)'],
            '原厂编号': ['原厂编号'],
            '主供应商': ['主供应商'],
            '商品级别': ['商品级别'],
            '大类': ['大类'],
            '中类': ['中类'],
            '子类': ['子类'],
            '内胆': ['内胆'],
            '年份': ['年份'],
            '季节': ['季节'],
            '尺寸组': ['尺寸组'],
            '品牌': ['品牌'],
            '渠道': ['渠道'],
            '退货仓': ['退货仓'],
            '预估采购价': ['预估采购价'],
            '标准价': ['标准价'],
            '面料成分': ['面料成分'],
            '里料成分': ['里料成分'],
            '其他成分': ['其他成分'],
            '执行标准': ['执行标准'],
            '安全类别': ['安全类别'],
            '备注': ['备注'],
        }
        self.load_workbook()

    def load_workbook(self):
        if self.wb:
            try:
                self.wb.close()
            except AttributeError:
                pass
        self.wb = load_workbook(self.file_path, data_only=True)

    def reload(self):
        self.load_workbook()
        if self.ws:
            self.set_sheet(self.ws.title)

    def close(self):
        if self.wb:
            try:
                self.wb.close()
            except AttributeError:
                pass
            finally:
                self.wb = None
        self.ws = None
        self.col_mapper = None
        self.cached_data.clear()

    def __del__(self):
        self.close()

    def get_allowed_sheets(self):
        return [s for s in self.wb.sheetnames if s in self.allowed_sheets]

    def set_sheet(self, sheet_name):
        if sheet_name not in self.wb.sheetnames:
            raise ValueError(f"工作表 {sheet_name} 不存在")
        self.ws = self.wb[sheet_name]
        self.col_mapper = ColumnMapper(self.ws)
        self.col_mapper.map_columns(self.keyword_map)
        self._build_cache()

    def _build_cache(self):
        self.cached_data.clear()
        max_row = self.ws.max_row
        max_col = self.ws.max_column
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                val = self.ws.cell(row, col).value
                self.cached_data[(row, col)] = val

    def get_cell_value(self, row, std_key):
        col = self.col_mapper.get_col(std_key)
        if col:
            val = self.cached_data.get((row, col))
            return str(val).strip() if val is not None else ""
        return ""

    def get_numeric_cell(self, row, std_key):
        col = self.col_mapper.get_col(std_key)
        if col:
            val = self.cached_data.get((row, col))
            if val is None:
                return None
            try:
                return float(val)
            except:
                return None
        return None

    def get_row_data_dict(self, row):
        d = {}
        for col, header in self.col_mapper.headers.items():
            val = self.cached_data.get((row, col))
            if isinstance(val, datetime):
                val = val.strftime("%Y-%m-%d %H:%M:%S")
            elif val is None:
                val = ""
            else:
                val = str(val).strip()
            d[header] = val
        return d

    def load_color_map(self):
        color_map = {}
        if 'CO' not in self.wb.sheetnames:
            return color_map
        co_ws = self.wb['CO']
        name_col = None
        code_col = None
        for col in range(1, co_ws.max_column + 1):
            header = co_ws.cell(1, col).value
            if not header:
                continue
            h = str(header).strip()
            if '颜色名称' in h or '颜色名' in h:
                name_col = col
            elif ('颜色' in h and '名称' not in h) or '颜色代码' in h:
                code_col = col
        if name_col and code_col:
            for row in range(2, co_ws.max_row + 1):
                name = co_ws.cell(row, name_col).value
                code = co_ws.cell(row, code_col).value
                if name and code:
                    color_map[str(name).strip()] = str(code).strip()
        return color_map

    def load_bz_standards(self):
        standards = set()
        if 'BZ' in self.wb.sheetnames:
            bz_ws = self.wb['BZ']
            for row in range(2, bz_ws.max_row + 1):
                val = bz_ws.cell(row, 2).value
                if val:
                    standards.add(str(val).strip())
        return standards